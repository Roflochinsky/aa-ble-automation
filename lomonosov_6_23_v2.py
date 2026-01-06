# 🚀 ТАЙМЛАЙНЫ AA_BLE v2 (окно 6:00–23:00, ПОМИНУТНО, SVG)
# - SVG-диаграмма с кнопками навигации < >
# - KPI-карточки: Зоны проведения работ / Зоны перерывов / Вне BLE / Прочее / Итого
# - Детализация по зонам и меткам (accordion, свёрнута по умолчанию)
# - Полностью оффлайн (без CDN)
print("🔄 Запуск анализа aa_ble v2 (6:00–23:00, поминутно, SVG-таймлайн)…")

# Для Google Colab
!pip install -q pandas openpyxl odfpy

import io, re, warnings, json
from datetime import datetime, timedelta, time as dtime
import pandas as pd
from google.colab import files
warnings.filterwarnings('ignore')

REEXPORT_CODE = "2025"
LAST_EXPORTS = []

def set_reexport_code(new_code):
    global REEXPORT_CODE
    REEXPORT_CODE = str(new_code)
    print("🔐 Код для повторной выгрузки обновлён.")

def list_saved_charts():
    if not LAST_EXPORTS:
        print("⚠️ Кэш пуст."); return
    print(f"📦 В кэше {len(LAST_EXPORTS)} диаграмм:")
    for i, it in enumerate(LAST_EXPORTS, 1):
        print(f"  {i}) дата={it['date_str']}, раздел={it['area_suffix']}, сотрудник={it.get('employee','')}")

def clear_saved_charts():
    LAST_EXPORTS.clear()
    print("🧹 Кэш диаграмм очищен.")

def round_051(x):
    if pd.isna(x): return x
    try: val = float(x)
    except: return x
    frac = abs(val) - abs(int(val))
    up = abs(int(val)) + 1; dn = abs(int(val))
    res = up if frac >= 0.51 else dn
    return res if val >= 0 else -res

def parse_date(s):
    if pd.isna(s): return None
    if isinstance(s, (pd.Timestamp, datetime)):
        return s.date()
    d = pd.to_datetime(s, yearfirst=True, errors='coerce')
    if pd.isna(d):
        d = pd.to_datetime(s, dayfirst=True, errors='coerce')
    return d.date() if pd.notna(d) else None

def parse_time_only(val):
    if pd.isna(val): return None
    s = str(val).strip()
    for fmt in ('%H:%M:%S','%H:%M'):
        try: return datetime.strptime(s, fmt).time()
        except: pass
    try:
        v = float(s); base = datetime(1899,12,30)
        return (base + timedelta(days=v)).time()
    except: return None

def parse_date_from_filename(filename: str):
    name = str(filename)
    m = re.search(r'(\d{4})[ _.-](\d{2})[ _.-](\d{2})', name)
    if m:
        try: return datetime(int(m[1]), int(m[2]), int(m[3])).date()
        except: pass
    m = re.search(r'(\d{2})[ _.-](\d{2})[ _.-](\d{4})', name)
    if m:
        try: return datetime(int(m[3]), int(m[2]), int(m[1])).date()
        except: pass
    return None

def infer_file_date_from_df(df):
    try:
        cols = [str(c).strip() for c in df.columns]
        low = {i: str(c).strip().lower() for i, c in enumerate(cols)}
        def find_col(candidates):
            for i, name in low.items():
                for cand in candidates:
                    if name == cand or cand in name:
                        return i
            return None
        ix_date = find_col(['date','дата'])
        ix_date_obj = find_col(['дата на объекте','date on site'])
        ix_shift = find_col(['shift day','shift_day','день смены','день'])
        for ix in [ix_date, ix_date_obj, ix_shift]:
            if ix is not None and ix < df.shape[1]:
                s = pd.to_datetime(df.iloc[:, ix], errors='coerce')
                val = s.dropna().iloc[0] if not s.dropna().empty else None
                if pd.notna(val):
                    return val.date()
        if df.shape[0] > 0:
            first_row = df.iloc[0, :]
            for i, v in enumerate(first_row.values):
                dt = pd.to_datetime(v, errors='coerce')
                if pd.notna(dt):
                    return dt.date()
        return None
    except:
        return None

def format_file_date_for_title(d):
    if not d: return ""
    return datetime.strftime(pd.to_datetime(d), "%d-%m-%Y")

def load_aable_second_sheet_multi(date_from=None, date_to=None):
    print("📈 Загрузите один или несколько файлов aa_ble (используется 2-й лист):")
    uploaded = files.upload()
    if not uploaded: return None
    frames = []
    for filename, raw in uploaded.items():
        try:
            if filename.endswith(('.xlsx','.xls')):
                xls = pd.ExcelFile(io.BytesIO(raw))
                sheet = xls.sheet_names[1] if len(xls.sheet_names) >= 2 else xls.sheet_names[0]
                df = xls.parse(sheet)
            elif filename.endswith('.csv'):
                df = pd.read_csv(io.StringIO(raw.decode('utf-8')))
            else:
                print(f"❌ Пропущен '{filename}' — неподдерживаемый формат"); continue
            df['__source__'] = filename
            inferred = infer_file_date_from_df(df)
            file_date = inferred or parse_date_from_filename(filename)
            df['__file_date__'] = file_date
            frames.append(df)
        except Exception as e:
            print(f"❌ Ошибка '{filename}': {e}")
    if not frames: return None
    all_df_raw = pd.concat(frames, ignore_index=True)
    print("🔎 Исходные колонки:", list(all_df_raw.columns))
    all_df = normalize_aable_columns(all_df_raw)
    all_df['shift_day'] = all_df['shift_day'].apply(parse_date)
    all_df['time_only'] = all_df['time_only'].apply(parse_time_only)
    print(f"🔍 После парсинга: записей={len(all_df)}")
    if date_from or date_to:
        dfrom = pd.to_datetime(date_from, dayfirst=True, errors='coerce').date() if date_from else None
        dto = pd.to_datetime(date_to, dayfirst=True, errors='coerce').date() if date_to else None
        mask = pd.Series(True, index=all_df.index)
        if dfrom: mask &= (all_df['shift_day'] >= dfrom)
        if dto:   mask &= (all_df['shift_day'] <= dto)
        all_df = all_df[mask]
    if '__file_date__' in all_df.columns:
        fd_series = pd.to_datetime(all_df['__file_date__'], errors='coerce').dt.date
        all_df = all_df[all_df['shift_day'] == fd_series]
    print(f"✅ aa_ble: записей={len(all_df)}")
    return all_df

def load_ble_journal():
    print("📖 Загрузите 'Журнал BLE' (метка → описание/локация):")
    uploaded = files.upload()
    if not uploaded: return None
    filename = list(uploaded.keys())[0]
    try:
        if filename.endswith(('.xlsx','.xls')):
            df = pd.read_excel(io.BytesIO(uploaded[filename]))
        elif filename.endswith('.csv'):
            df = pd.read_csv(io.StringIO(uploaded[filename].decode('utf-8')))
        else:
            print("❌ Неподдерживаемый формат журнала BLE"); return None
        print(f"✅ Журнал BLE загружен: {len(df)} записей")
        return df
    except Exception as e:
        print(f"❌ Ошибка чтения журнала BLE: {e}")
        return None

def load_people_mapping():
    print("🗂 Загрузите 'Привязка людей' (A: ТН, B: ФИО, D: Участок Работ):")
    uploaded = files.upload()
    if not uploaded: return None
    filename = list(uploaded.keys())[0]
    try:
        if filename.endswith(('.xlsx','.xls')):
            df = pd.read_excel(io.BytesIO(uploaded[filename]))
        elif filename.endswith('.csv'):
            df = pd.read_csv(io.StringIO(uploaded[filename].decode('utf-8')))
        else:
            print("❌ Неподдерживаемый формат файла привязки"); return None
        print(f"✅ Привязка людей загружена: {len(df)} записей")
        return df
    except Exception as e:
        print(f"❌ Ошибка чтения файла привязки: {e}")
        return None

def build_area_map(mapping_df):
    area_map = {}
    if mapping_df is None or mapping_df.empty: return area_map
    for _, row in mapping_df.iterrows():
        try:
            tn = str(row.iloc[0]).strip()
            area = str(row.iloc[3]).strip() if len(row) > 3 else ""
            if tn and area and area.lower() != 'nan':
                area_map[tn] = area
        except: continue
    print(f"✅ Сформирована карта участков: {len(area_map)} сотрудников")
    return area_map

def build_fio_map(mapping_df):
    fio_map = {}
    if mapping_df is None or mapping_df.empty: return fio_map
    for _, row in mapping_df.iterrows():
        try:
            tn = str(row.iloc[0]).strip()
            fio = str(row.iloc[1]).strip() if len(row) > 1 else ""
            if tn and fio and fio.lower() != 'nan':
                fio_map[tn] = fio
        except: continue
    print(f"✅ Сформирована карта ФИО: {len(fio_map)} сотрудников")
    return fio_map

def normalize_aable_columns(df):
    orig = df.copy()
    cols = list(orig.columns)
    low = {i: str(c).strip().lower() for i, c in enumerate(cols)}
    def find_by_name(needles):
        for i, name in low.items():
            for n in needles:
                if name == n: return i
        for i, name in low.items():
            for n in needles:
                if n in name: return i
        return None
    POS = {'tn_number':0,'shift_day':3,'ble_tag':10,'zone_id':11,'time_only':15}
    IX_TN   = find_by_name(['тн','tn','таб','табель','tabnum','tab_num','personal'])
    IX_DAY  = find_by_name(['день смены','shift day','shift_day','день','смены'])
    IX_TAG  = find_by_name(['ble','метк','mekta','metka','маяч','tag'])
    IX_ZONE = find_by_name(['zona','зона','zone'])
    IX_TIME = find_by_name(['время на объекте','время','time'])
    def pick(ix_name, pos_key):
        ix = {'tn_number':IX_TN,'shift_day':IX_DAY,'ble_tag':IX_TAG,'zone_id':IX_ZONE,'time_only':IX_TIME}[ix_name]
        if ix is not None and ix < orig.shape[1]: return orig.iloc[:, ix]
        pos = POS[pos_key]
        return orig.iloc[:, pos] if pos < orig.shape[1] else pd.Series([None]*len(orig))
    out = pd.DataFrame({
        'tn_number': pick('tn_number','tn_number'),
        'shift_day': pick('shift_day','shift_day'),
        'ble_tag': pick('ble_tag','ble_tag'),
        'zone_id': pick('zone_id','zone_id'),
        'time_only': pick('time_only','time_only'),
    })
    if '__source__' in orig.columns: out['__source__'] = orig['__source__']
    if '__file_date__' in orig.columns: out['__file_date__'] = orig['__file_date__']
    return out

def attach_ble_desc_map(ble_journal_df):
    if ble_journal_df is None or ble_journal_df.empty: return {}
    jr = ble_journal_df.copy()
    jr.iloc[:,0] = jr.iloc[:,0].astype(str).str.strip()
    if jr.shape[1] < 4: return {}
    jr['__desc__'] = jr.iloc[:,3].astype(str).str.strip()
    d = {}
    for _, r in jr.iterrows():
        key = r.iloc[0]; val = r['__desc__']
        if str(val).strip(): d[str(key)] = val
    return d

# Константы
ZONE_COLORS = {
    0: '#95A5A6', 1: '#27AE60', 2: '#E67E22', 3: '#E74C3C',
    4: '#E74C3C', 5: '#8E44AD', 6: '#3498DB', 7: '#F1C40F',
    8: '#1ABC9C', 9: '#9B59B6', 10: '#1F77B4', 11: '#E67E22',
    12: '#2ECC71', 13: '#E67E22'
}
ZONE_NAMES = {
    0: 'Вне зоны BLE-маячков', 1: 'Зоны проведения работ', 2: 'Столовые',
    3: 'Опасные зоны', 4: 'Курилки', 5: 'Зоны отдыха', 6: 'ВЖГ',
    7: 'Туалеты', 8: 'Остановки автобусов', 9: 'Административные помещения',
    10: 'Зона выдачи WW', 11: 'Склад', 12: 'Мастерские', 13: 'КПП'
}
ZONE_ORDER = [1, 4, 10, 13, 2, 7, 0, 5, 3, 6, 8, 9, 11, 12]

# KPI группировка
KPI_GROUPS = {
    'work': [1],
    'breaks': [2, 4, 5, 7],
    'no_ble': [0],
    'other': [3, 6, 8, 9, 10, 11, 12, 13]
}

def build_user_minute_bins_by_tag(df_user):
    rows = df_user.dropna(subset=['shift_day','time_only']).sort_values(['shift_day','time_only']).to_dict('records')
    if not rows: return []
    minute_bins = []; prev_dt = None; current_day = rows[0]['shift_day']
    for r in rows:
        day = r['shift_day'] or current_day
        t = r['time_only']
        if t is None: continue
        dt = datetime.combine(day, t)
        if prev_dt and dt < prev_dt:
            dt = datetime.combine(day + timedelta(days=1), t)
            current_day = day + timedelta(days=1)
        else:
            current_day = day
        try: tag_val = int(float(str(r.get('ble_tag')))) if r.get('ble_tag') not in [None,""] else 0
        except: tag_val = 0
        try: zone_val = int(r.get('zone_id')) if pd.notna(r.get('zone_id')) else 0
        except: zone_val = 0
        minute_bins.append({'start': dt, 'end': dt + timedelta(minutes=1), 'ble_tag': tag_val, 'zone_id': zone_val})
        prev_dt = dt
    return minute_bins

def build_segments_all(df, area_map, fio_map):
    out_rows = []
    for (tn, shift_day), g in df.groupby(['tn_number','shift_day']):
        bins = build_user_minute_bins_by_tag(g)
        file_dates = g['__file_date__'].dropna().astype(str)
        file_date = file_dates.mode().iloc[0] if not file_dates.empty else None
        for b in bins:
            area = area_map.get(str(tn).strip(), '')
            fio = fio_map.get(str(tn).strip(), f"ТН {str(tn).strip()}")
            zid = b['zone_id'] if pd.notna(b['zone_id']) else 0
            out_rows.append({
                'tn_number': str(tn).strip(),
                'employee': fio,
                'area': area,
                'date': b['start'].date(),
                'file_date': file_date,
                'start': b['start'],
                'end': b['end'],
                'duration_minutes': 1.0,
                'ble_tag': b['ble_tag'],
                'zone_id': zid,
                'zone_name': ZONE_NAMES.get(zid, f"Зона {zid}"),
            })
    out = pd.DataFrame(out_rows)
    if not out.empty:
        out = out.sort_values(['date','employee','start'])
        out['duration_minutes'] = out['duration_minutes'].apply(round_051)
    return out

def lighten_color(hex_color, factor=0.4):
    hex_color = hex_color.lstrip('#')
    r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
    r = int(r + (255 - r) * factor)
    g = int(g + (255 - g) * factor)
    b = int(b + (255 - b) * factor)
    return f'#{r:02x}{g:02x}{b:02x}'

def format_duration(mins):
    h = int(mins) // 60
    m = int(mins) % 60
    if h == 0: return f"{m} мин"
    return f"{h}ч {m:02d}м"

def prepare_timeline_data(df_user, tag_desc_map):
    """Подготавливает данные для SVG-таймлайна (24ч, динамический viewport)"""
    if df_user is None or df_user.empty:
        return None
    
    df = df_user.copy()
    df['start'] = pd.to_datetime(df['start'])
    df['end'] = pd.to_datetime(df['end'])
    
    if df.empty:
        return None
    
    user = df['employee'].iloc[0]
    area = df['area'].iloc[0] if 'area' in df.columns else ''
    
    # Определяем базовую дату (начало суток) — берём дату первой записи
    first_dt = df['start'].min()
    base_date = first_dt.date()
    base_midnight = datetime.combine(base_date, dtime(0, 0))
    
    # Для ночных смен: если есть записи после полуночи следующего дня, 
    # они будут иметь minute > 1440
    
    # Собираем сегменты с абсолютными минутами от полуночи base_date
    segments = []
    for _, row in df.iterrows():
        # Минуты от базовой полуночи (может быть > 1440 для ночных смен)
        minute = int((row['start'] - base_midnight).total_seconds() // 60)
        zone = int(row['zone_id']) if pd.notna(row['zone_id']) else 0
        tag = int(row['ble_tag']) if pd.notna(row['ble_tag']) else 0
        desc = tag_desc_map.get(str(tag), '')
        segments.append({
            'minute': minute,
            'zone': zone,
            'tag': tag,
            'desc': desc,
            'datetime': row['start']
        })
    
    if not segments:
        return None
    
    # Динамический viewport: от первой до последней минуты
    all_minutes = [s['minute'] for s in segments]
    first_minute = min(all_minutes)
    last_minute = max(all_minutes)
    
    # Округляем до часов для viewport
    viewport_start = (first_minute // 60) * 60  # округление вниз до часа
    viewport_end = ((last_minute // 60) + 1) * 60  # округление вверх до часа
    
    # Статистика по зонам
    zone_stats = {}
    for seg in segments:
        z = seg['zone']
        if z not in zone_stats:
            zone_stats[z] = {'minutes': 0, 'tags': {}}
        zone_stats[z]['minutes'] += 1
        t = seg['tag']
        if t not in zone_stats[z]['tags']:
            zone_stats[z]['tags'][t] = {'minutes': 0, 'desc': seg['desc']}
        zone_stats[z]['tags'][t]['minutes'] += 1
    
    # KPI — только за фактическое время присутствия
    total = len(segments)
    kpi = {'work': 0, 'breaks': 0, 'no_ble': 0, 'other': 0}
    for seg in segments:
        z = seg['zone']
        if z in KPI_GROUPS['work']: kpi['work'] += 1
        elif z in KPI_GROUPS['breaks']: kpi['breaks'] += 1
        elif z in KPI_GROUPS['no_ble']: kpi['no_ble'] += 1
        else: kpi['other'] += 1
    
    return {
        'employee': user,
        'area': area,
        'base_date': base_date,
        'segments': segments,
        'zone_stats': zone_stats,
        'kpi': kpi,
        'total': total,
        'viewport_start': viewport_start,  # минута начала viewport (округлено до часа)
        'viewport_end': viewport_end,      # минута конца viewport (округлено до часа)
        'first_minute': first_minute,      # первая фактическая минута
        'last_minute': last_minute,        # последняя фактическая минута
    }

def generate_svg_timeline_html(data, tag_desc_map):
    """Генерирует полный HTML с SVG-таймлайном для одного сотрудника"""
    if data is None:
        return None
    
    employee = data['employee']
    area = data['area']
    base_date = data['base_date']
    segments = data['segments']
    zone_stats = data['zone_stats']
    kpi = data['kpi']
    total = data['total']
    viewport_start = data['viewport_start']
    viewport_end = data['viewport_end']
    
    # Форматирование даты для заголовка
    date_str = base_date.strftime('%d.%m.%Y') if base_date else ''
    
    # Расчёт процентов
    def pct(v): return round(v / total * 100) if total > 0 else 0
    
    # Генерация SVG сегментов (позиция относительно viewport_start)
    svg_rects = []
    for seg in segments:
        # Позиция в SVG = (абсолютная минута - начало viewport) * ширина блока
        x_pos = (seg['minute'] - viewport_start) * 20
        color = ZONE_COLORS.get(seg['zone'], '#95A5A6')
        svg_rects.append(
            f'<rect x="{x_pos}" y="0" width="19" height="100" '
            f'fill="{color}" rx="3" class="segment" '
            f'data-minute="{seg["minute"]}" data-zone="{seg["zone"]}" '
            f'data-tag="{seg["tag"]}" data-desc="{seg["desc"]}"/>'
        )
    svg_content = '\n'.join(svg_rects)
    
    # Ширина SVG = (конец viewport - начало viewport) * 20
    svg_width = (viewport_end - viewport_start) * 20
    
    # Генерация accordion зон
    zones_html = []
    sorted_zones = sorted(zone_stats.keys(), key=lambda z: ZONE_ORDER.index(z) if z in ZONE_ORDER else 999)
    for zid in sorted_zones:
        zdata = zone_stats[zid]
        mins = zdata['minutes']
        zpct = round(mins / total * 100, 1) if total > 0 else 0
        color = ZONE_COLORS.get(zid, '#95A5A6')
        light_color = lighten_color(color)
        zname = ZONE_NAMES.get(zid, f'Зона {zid}')
        
        # Метки внутри зоны
        tags_html = []
        sorted_tags = sorted(zdata['tags'].items(), key=lambda x: x[1]['minutes'], reverse=True)
        for tag_id, tdata in sorted_tags:
            tmins = tdata['minutes']
            tpct = round(tmins / total * 100, 1) if total > 0 else 0
            tdesc = tdata['desc']
            tlabel = f"#{tag_id} ({tdesc})" if tdesc else f"#{tag_id}"
            tags_html.append(
                f'<div class="tag-row" style="background:{light_color}">'
                f'<span class="tag-name">{tlabel}</span>'
                f'<span class="tag-value">{format_duration(tmins)} ({tpct}%)</span></div>'
            )
        
        zones_html.append(f'''
        <div class="zone-accordion">
          <div class="zone-header" onclick="toggleZone(this)">
            <span class="zone-toggle">▶</span>
            <div class="zone-dot" style="background:{color}"></div>
            <span class="zone-name">{zname}</span>
            <span class="zone-value">{format_duration(mins)} ({zpct}%)</span>
          </div>
          <div class="zone-content">{''.join(tags_html)}</div>
        </div>''')
    
    # Собираем JSON для JS (без datetime, только нужные поля)
    segments_for_js = [{'minute': s['minute'], 'zone': s['zone'], 'tag': s['tag'], 'desc': s['desc']} for s in segments]
    js_data = json.dumps({
        'segments': segments_for_js,
        'zoneColors': ZONE_COLORS,
        'zoneNames': ZONE_NAMES,
        'viewportStart': viewport_start,
        'viewportEnd': viewport_end,
        'baseDate': date_str,
    }, ensure_ascii=False)

    # HTML шаблон (без даты в subtitle)
    html = f'''<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="utf-8">
<title>AA_BLE Таймлайн — {employee}</title>
<style>
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{ font-family: 'Segoe UI', Arial, sans-serif; background: #f0f2f5; padding: 24px; color: #333; }}
.container {{ max-width: 1400px; margin: 0 auto; }}
.card {{ background: #fff; border-radius: 12px; box-shadow: 0 2px 12px rgba(0,0,0,0.08); padding: 20px; margin-bottom: 20px; }}
.header {{ display: flex; justify-content: space-between; align-items: flex-start; margin-bottom: 20px; }}
.header-left {{ flex-shrink: 0; }}
.title {{ font-size: 22px; font-weight: 600; color: #1a1a2e; }}
.subtitle {{ font-size: 14px; color: #666; margin-top: 4px; }}
.time-display {{ background: #1976d2; color: #fff; padding: 8px 16px; border-radius: 8px; font-weight: 600; font-size: 15px; margin-top: 10px; display: inline-block; }}
.kpi-section {{ margin-bottom: 20px; }}
.kpi-cards {{ display: flex; gap: 12px; flex-wrap: wrap; }}
.kpi-card {{ flex: 1; min-width: 140px; padding: 16px; border-radius: 10px; text-align: center; position: relative; overflow: hidden; }}
.kpi-card::before {{ content: ''; position: absolute; left: 0; top: 0; bottom: 0; width: 4px; }}
.kpi-card.work {{ background: #e8f5e9; }} .kpi-card.work::before {{ background: #27AE60; }}
.kpi-card.breaks {{ background: #ffebee; }} .kpi-card.breaks::before {{ background: #E74C3C; }}
.kpi-card.no-ble {{ background: #f5f5f5; }} .kpi-card.no-ble::before {{ background: #95A5A6; }}
.kpi-card.other {{ background: #fff3e0; }} .kpi-card.other::before {{ background: #E67E22; }}
.kpi-card.total {{ background: #e3f2fd; }} .kpi-card.total::before {{ background: #1976d2; }}
.kpi-label {{ font-size: 12px; color: #666; margin-bottom: 4px; text-transform: uppercase; letter-spacing: 0.5px; }}
.kpi-value {{ font-size: 20px; font-weight: 700; color: #1a1a2e; }}
.kpi-pct {{ font-size: 14px; color: #666; margin-top: 2px; }}
.timeline-wrapper {{ position: relative; display: flex; align-items: center; gap: 12px; margin-top: 20px; }}
.nav-btn {{ width: 48px; height: 100px; border: none; border-radius: 8px; cursor: pointer; font-size: 24px; font-weight: bold; transition: all 0.2s; display: flex; align-items: center; justify-content: center; flex-shrink: 0; background: #e8f4fd; color: #1976d2; }}
.nav-btn:hover {{ background: #bbdefb; }}
.nav-btn:active {{ transform: scale(0.95); background: #90caf9; }}
.nav-btn:disabled {{ background: #f5f5f5; color: #ccc; cursor: not-allowed; transform: none; }}
.viewport {{ flex: 1; overflow: hidden; border-radius: 8px; background: #fafafa; border: 1px solid #e0e0e0; }}
.canvas-container {{ transition: transform 0.3s cubic-bezier(0.4, 0, 0.2, 1); }}
.segment {{ cursor: pointer; transition: filter 0.15s; }}
.segment:hover {{ filter: brightness(1.15) saturate(1.1); }}
.time-axis-wrapper {{ margin-left: 60px; margin-right: 60px; overflow: hidden; }}
.time-axis {{ display: flex; margin-top: 8px; position: relative; height: 24px; }}
.time-axis-inner {{ display: flex; transition: transform 0.3s cubic-bezier(0.4, 0, 0.2, 1); }}
.time-label {{ font-size: 11px; color: #666; text-align: left; flex-shrink: 0; width: 300px; }}
</style>
</head>
<body>
<div class="container">
<div class="card">
<div class="header">
  <div class="header-left">
    <div class="title">{employee}</div>
    <div class="subtitle">Участок: {area if area else 'Не указан'}</div>
    <div class="time-display" id="timeRange">{date_str} 00:00 — 01:00</div>
  </div>
</div>
<div class="kpi-section">
  <div class="kpi-cards">
    <div class="kpi-card work"><div class="kpi-label">Зоны проведения работ</div><div class="kpi-value">{format_duration(kpi['work'])}</div><div class="kpi-pct">{pct(kpi['work'])}%</div></div>
    <div class="kpi-card breaks"><div class="kpi-label">Зоны перерывов</div><div class="kpi-value">{format_duration(kpi['breaks'])}</div><div class="kpi-pct">{pct(kpi['breaks'])}%</div></div>
    <div class="kpi-card no-ble"><div class="kpi-label">Вне BLE</div><div class="kpi-value">{format_duration(kpi['no_ble'])}</div><div class="kpi-pct">{pct(kpi['no_ble'])}%</div></div>
    <div class="kpi-card other"><div class="kpi-label">Прочее</div><div class="kpi-value">{format_duration(kpi['other'])}</div><div class="kpi-pct">{pct(kpi['other'])}%</div></div>
    <div class="kpi-card total"><div class="kpi-label">Итого</div><div class="kpi-value">{format_duration(total)}</div><div class="kpi-pct">100%</div></div>
  </div>
</div>
'''
    return html, svg_content, zones_html, js_data, svg_width, viewport_start, viewport_end

def generate_full_html(data, tag_desc_map):
    """Генерирует полный HTML файл"""
    result = generate_svg_timeline_html(data, tag_desc_map)
    if result is None:
        return None
    
    html_start, svg_content, zones_html, js_data, svg_width, viewport_start, viewport_end = result
    
    html = html_start + f'''
<div class="timeline-wrapper">
  <button class="nav-btn" id="btnLeft" onclick="scrollTimeline(-60)">&lt;</button>
  <div class="viewport">
    <div class="canvas-container" id="canvasContainer">
      <svg id="timelineSvg" width="{svg_width}" height="100">
        {svg_content}
      </svg>
    </div>
  </div>
  <button class="nav-btn" id="btnRight" onclick="scrollTimeline(60)">&gt;</button>
</div>
<div class="time-axis-wrapper">
  <div class="time-axis">
    <div class="time-axis-inner" id="timeAxis"></div>
  </div>
</div>
<style>
.details-section {{ margin-top: 20px; border-top: 1px solid #eee; padding-top: 16px; }}
.details-toggle {{ display: flex; align-items: center; gap: 8px; cursor: pointer; padding: 10px; border-radius: 8px; font-weight: 500; color: #555; transition: background 0.2s; }}
.details-toggle:hover {{ background: #f5f5f5; }}
.details-icon {{ font-size: 12px; transition: transform 0.2s; }}
.details-icon.open {{ transform: rotate(90deg); }}
.details-content {{ margin-top: 12px; display: none; }}
.zone-accordion {{ margin-bottom: 2px; }}
.zone-header {{ display: flex; align-items: center; padding: 8px 10px; cursor: pointer; border-radius: 6px; transition: background 0.2s; font-size: 13px; }}
.zone-header:hover {{ background: #f5f5f5; }}
.zone-toggle {{ font-size: 10px; margin-right: 8px; color: #666; transition: transform 0.2s; width: 12px; }}
.zone-toggle.open {{ transform: rotate(90deg); }}
.zone-dot {{ width: 12px; height: 12px; border-radius: 4px; margin-right: 10px; flex-shrink: 0; }}
.zone-name {{ flex: 1; color: #333; }}
.zone-value {{ font-weight: 600; color: #333; margin-left: 8px; white-space: nowrap; }}
.zone-content {{ margin-left: 30px; padding-left: 10px; border-left: 2px solid #e0e0e0; display: none; }}
.tag-row {{ display: flex; align-items: center; justify-content: space-between; padding: 6px 10px; margin: 3px 0; border-radius: 6px; font-size: 12px; }}
.tag-name {{ flex: 1; color: #555; }}
.tag-value {{ font-weight: 600; color: #333; margin-left: 8px; }}
.tooltip {{ position: fixed; background: #1a1a2e; color: #fff; padding: 12px 16px; border-radius: 8px; font-size: 13px; pointer-events: none; opacity: 0; transition: opacity 0.2s; z-index: 1000; max-width: 280px; box-shadow: 0 4px 20px rgba(0,0,0,0.25); }}
.tooltip.visible {{ opacity: 1; }}
.tooltip-row {{ margin: 4px 0; display: flex; justify-content: space-between; gap: 16px; }}
.tooltip-label {{ color: #aaa; }}
.tooltip-value {{ font-weight: 600; }}
.tooltip-header {{ font-weight: 600; font-size: 14px; margin-bottom: 8px; padding-bottom: 8px; border-bottom: 1px solid #333; }}
</style>
<div class="details-section">
  <div class="details-toggle" onclick="toggleDetails()">
    <span class="details-icon" id="detailsIcon">▶</span>
    <span>Детализация по зонам и меткам</span>
  </div>
  <div class="details-content" id="detailsContent">
    {''.join(zones_html)}
  </div>
</div>
</div>
</div>
<div class="tooltip" id="tooltip"></div>
'''
    return html, js_data, viewport_start, viewport_end

def generate_js_script(js_data, viewport_start, viewport_end):
    """Генерирует JavaScript для интерактивности с динамическим viewport"""
    total_minutes = viewport_end - viewport_start
    
    return f'''
<script>
const DATA = {js_data};
const CONFIG = {{
  MINUTE_WIDTH: 20,
  SEGMENT_HEIGHT: 100,
  SCROLL_STEP: 60,
  WINDOW_START: {viewport_start},
  WINDOW_END: {viewport_end},
}};
const TOTAL_MINUTES = CONFIG.WINDOW_END - CONFIG.WINDOW_START;
let offset = 0;
let viewportMinutes = 60; // будет пересчитано при загрузке

function calcViewportMinutes() {{
  const viewport = document.querySelector('.viewport');
  if (viewport) {{
    viewportMinutes = Math.floor(viewport.clientWidth / CONFIG.MINUTE_WIDTH);
  }}
}}

// Форматирование времени с учётом перехода через полночь
function formatTime(m) {{
  // m может быть > 1440 для ночных смен
  const h = Math.floor(m / 60) % 24;
  const mm = m % 60;
  return h.toString().padStart(2, '0') + ':' + mm.toString().padStart(2, '0');
}}

// Форматирование даты+времени для time-display
function formatDateTime(m) {{
  const baseDate = DATA.baseDate || '';
  // Определяем, перешли ли через полночь (m >= 1440)
  const dayOffset = Math.floor(m / 1440);
  const timeStr = formatTime(m);
  
  if (dayOffset > 0 && baseDate) {{
    // Парсим базовую дату и добавляем дни
    const parts = baseDate.split('.');
    if (parts.length === 3) {{
      const d = new Date(parts[2], parts[1] - 1, parseInt(parts[0]) + dayOffset);
      const newDate = d.getDate().toString().padStart(2, '0') + '.' + 
                      (d.getMonth() + 1).toString().padStart(2, '0') + '.' + 
                      d.getFullYear();
      return newDate + ' ' + timeStr;
    }}
  }}
  return baseDate + ' ' + timeStr;
}}

function renderTimeAxis() {{
  const axis = document.getElementById('timeAxis');
  axis.innerHTML = '';
  for (let m = 0; m <= TOTAL_MINUTES; m += 15) {{
    const label = document.createElement('div');
    label.className = 'time-label';
    label.style.width = (15 * CONFIG.MINUTE_WIDTH) + 'px';
    label.textContent = formatTime(CONFIG.WINDOW_START + m);
    axis.appendChild(label);
  }}
}}

function scrollTimeline(minutes) {{
  calcViewportMinutes();
  const maxOffset = Math.max(0, TOTAL_MINUTES - viewportMinutes);
  offset = Math.max(0, Math.min(offset + minutes, maxOffset));
  updateOffset();
}}

function updateOffset() {{
  calcViewportMinutes();
  const px = offset * CONFIG.MINUTE_WIDTH;
  document.getElementById('canvasContainer').style.transform = 'translateX(-' + px + 'px)';
  document.getElementById('timeAxis').style.transform = 'translateX(-' + px + 'px)';
  
  const maxOffset = Math.max(0, TOTAL_MINUTES - viewportMinutes);
  document.getElementById('btnLeft').disabled = (offset <= 0);
  document.getElementById('btnRight').disabled = (offset >= maxOffset);
  
  const startMins = CONFIG.WINDOW_START + offset;
  const endMins = startMins + 60;
  document.getElementById('timeRange').textContent = formatDateTime(startMins) + ' — ' + formatTime(endMins);
}}

function toggleDetails() {{
  const content = document.getElementById('detailsContent');
  const icon = document.getElementById('detailsIcon');
  if (content.style.display === 'none' || content.style.display === '') {{
    content.style.display = 'block';
    icon.classList.add('open');
  }} else {{
    content.style.display = 'none';
    icon.classList.remove('open');
  }}
}}

function toggleZone(header) {{
  const content = header.nextElementSibling;
  const toggle = header.querySelector('.zone-toggle');
  if (content.style.display === 'none' || content.style.display === '') {{
    content.style.display = 'block';
    toggle.classList.add('open');
  }} else {{
    content.style.display = 'none';
    toggle.classList.remove('open');
  }}
}}

function showTooltip(e) {{
  const tooltip = document.getElementById('tooltip');
  const rect = e.target;
  const minute = parseInt(rect.dataset.minute);
  const timeStr = formatDateTime(minute);
  const zone = parseInt(rect.dataset.zone);
  const tag = rect.dataset.tag;
  const desc = rect.dataset.desc || '';
  tooltip.innerHTML = '<div class="tooltip-header">' + timeStr + '</div>' +
    '<div class="tooltip-row"><span class="tooltip-label">Зона:</span><span class="tooltip-value">' + (DATA.zoneNames[zone] || 'Зона ' + zone) + '</span></div>' +
    '<div class="tooltip-row"><span class="tooltip-label">Метка:</span><span class="tooltip-value">#' + tag + '</span></div>' +
    '<div class="tooltip-row"><span class="tooltip-label">Описание:</span><span class="tooltip-value">' + desc + '</span></div>';
  tooltip.classList.add('visible');
  moveTooltip(e);
}}

function moveTooltip(e) {{
  const tooltip = document.getElementById('tooltip');
  const tooltipWidth = tooltip.offsetWidth || 280;
  const tooltipHeight = tooltip.offsetHeight || 150;
  const windowWidth = window.innerWidth;
  const windowHeight = window.innerHeight;
  
  let left = e.clientX + 15;
  let top = e.clientY + 15;
  
  // Проверяем правую границу
  if (left + tooltipWidth > windowWidth - 10) {{
    left = e.clientX - tooltipWidth - 15;
  }}
  // Проверяем нижнюю границу
  if (top + tooltipHeight > windowHeight - 10) {{
    top = e.clientY - tooltipHeight - 15;
  }}
  // Не уходим за левый край
  if (left < 10) left = 10;
  // Не уходим за верхний край
  if (top < 10) top = 10;
  
  tooltip.style.left = left + 'px';
  tooltip.style.top = top + 'px';
}}

function hideTooltip() {{
  document.getElementById('tooltip').classList.remove('visible');
}}

document.addEventListener('DOMContentLoaded', function() {{
  calcViewportMinutes();
  renderTimeAxis();
  updateOffset();
  document.querySelectorAll('.segment').forEach(function(el) {{
    el.addEventListener('mouseenter', showTooltip);
    el.addEventListener('mouseleave', hideTooltip);
    el.addEventListener('mousemove', moveTooltip);
  }});
}});

// Пересчитываем при изменении размера окна
window.addEventListener('resize', function() {{
  calcViewportMinutes();
  updateOffset();
}});

document.addEventListener('keydown', function(e) {{
  if (e.key === 'ArrowLeft') scrollTimeline(-CONFIG.SCROLL_STEP);
  if (e.key === 'ArrowRight') scrollTimeline(CONFIG.SCROLL_STEP);
}});
</script>
</body>
</html>
'''

def export_svg_timeline(data, tag_desc_map, filename):
    """Экспортирует SVG-таймлайн в HTML файл"""
    result = generate_full_html(data, tag_desc_map)
    if result is None:
        print(f"⚠️ Нет данных для {filename}")
        return
    
    html_content, js_data, viewport_start, viewport_end = result
    js_script = generate_js_script(js_data, viewport_start, viewport_end)
    full_html = html_content + js_script
    
    with open(filename, 'w', encoding='utf-8') as f:
        f.write(full_html)
    files.download(filename)
    print(f"✅ HTML '{filename}' скачан")

def export_combined_svg_html(all_data, tag_desc_map, filename, page_title):
    """Экспортирует все таймлайны в один HTML файл с оглавлением"""
    if not all_data:
        print("⚠️ Нет данных для экспорта")
        return
    
    # Оглавление
    toc_items = []
    sections = []
    
    for i, data in enumerate(all_data, 1):
        anchor = f"sec{i}"
        employee = data['employee']
        base_date = data.get('base_date')
        date_str = base_date.strftime('%d.%m.%Y') if base_date else ''
        toc_items.append(f'<li><a href="#{anchor}">{date_str} — {employee}</a></li>')
        
        result = generate_full_html(data, tag_desc_map)
        if result is None:
            continue
        html_content, js_data, viewport_start, viewport_end = result
        
        # Модифицируем ID элементов чтобы не было конфликтов
        html_content = html_content.replace('id="timeRange"', f'id="timeRange{i}"')
        html_content = html_content.replace('id="canvasContainer"', f'id="canvasContainer{i}"')
        html_content = html_content.replace('id="timeAxis"', f'id="timeAxis{i}"')
        html_content = html_content.replace('id="btnLeft"', f'id="btnLeft{i}"')
        html_content = html_content.replace('id="btnRight"', f'id="btnRight{i}"')
        html_content = html_content.replace('id="detailsContent"', f'id="detailsContent{i}"')
        html_content = html_content.replace('id="detailsIcon"', f'id="detailsIcon{i}"')
        html_content = html_content.replace('id="timelineSvg"', f'id="timelineSvg{i}"')
        html_content = html_content.replace('id="tooltip"', f'id="tooltip{i}"')
        html_content = html_content.replace('onclick="scrollTimeline(-60)"', f'onclick="scrollTimeline{i}(-60)"')
        html_content = html_content.replace('onclick="scrollTimeline(60)"', f'onclick="scrollTimeline{i}(60)"')
        html_content = html_content.replace('onclick="toggleDetails()"', f'onclick="toggleDetails{i}()"')
        html_content = html_content.replace('onclick="toggleZone(this)"', f'onclick="toggleZone{i}(this)"')
        
        sections.append({
            'anchor': anchor,
            'html': html_content,
            'js_data': js_data,
            'index': i,
            'viewport_start': viewport_start,
            'viewport_end': viewport_end,
        })
    
    # Генерируем полный HTML
    combined_html = f'''<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="utf-8">
<title>{page_title}</title>
<style>
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{ font-family: 'Segoe UI', Arial, sans-serif; background: #f0f2f5; padding: 24px; color: #333; }}
.main-container {{ max-width: 1400px; margin: 0 auto; }}
.toc-card {{ background: #fff; border-radius: 12px; box-shadow: 0 2px 12px rgba(0,0,0,0.08); padding: 24px; margin-bottom: 24px; }}
.toc-title {{ font-size: 24px; font-weight: 600; color: #1a1a2e; margin-bottom: 16px; }}
.toc-list {{ list-style: none; }}
.toc-list li {{ margin: 8px 0; }}
.toc-list a {{ color: #1976d2; text-decoration: none; font-size: 15px; }}
.toc-list a:hover {{ text-decoration: underline; }}
.section {{ margin-bottom: 32px; }}
.back-link {{ margin-bottom: 12px; }}
.back-link a {{ color: #1976d2; text-decoration: none; font-size: 14px; }}
.back-link a:hover {{ text-decoration: underline; }}
</style>
</head>
<body>
<div class="main-container">
<div class="toc-card" id="toc">
  <div class="toc-title">{page_title}</div>
  <ul class="toc-list">{''.join(toc_items)}</ul>
</div>
'''
    
    # Добавляем секции
    for sec in sections:
        i = sec['index']
        combined_html += f'''
<div class="section" id="{sec['anchor']}">
  <div class="back-link"><a href="#toc">← К оглавлению</a></div>
  {sec['html']}
'''
        # Генерируем уникальный JS для этой секции
        combined_html += generate_section_js(sec['js_data'], i, sec['viewport_start'], sec['viewport_end'])
        combined_html += '</div>'
    
    combined_html += '''
</div>
</body>
</html>'''
    
    with open(filename, 'w', encoding='utf-8') as f:
        f.write(combined_html)
    files.download(filename)
    print(f"✅ Объединённый HTML '{filename}' скачан")

def generate_section_js(js_data, index, viewport_start, viewport_end):
    """Генерирует уникальный JS для секции с индексом и динамическим viewport"""
    total_minutes = viewport_end - viewport_start
    
    return f'''
<script>
(function() {{
  const DATA{index} = {js_data};
  const CONFIG{index} = {{
    MINUTE_WIDTH: 20,
    SEGMENT_HEIGHT: 100,
    SCROLL_STEP: 60,
    WINDOW_START: {viewport_start},
    WINDOW_END: {viewport_end},
  }};
  const TOTAL_MINUTES{index} = CONFIG{index}.WINDOW_END - CONFIG{index}.WINDOW_START;
  let offset{index} = 0;
  let viewportMinutes{index} = 60; // будет пересчитано при загрузке

  function calcViewportMinutes{index}() {{
    const viewport = document.querySelector('#sec{index} .viewport');
    if (viewport) {{
      viewportMinutes{index} = Math.floor(viewport.clientWidth / CONFIG{index}.MINUTE_WIDTH);
    }}
  }}

  function formatTime{index}(m) {{
    const h = Math.floor(m / 60) % 24;
    const mm = m % 60;
    return h.toString().padStart(2, '0') + ':' + mm.toString().padStart(2, '0');
  }}

  function formatDateTime{index}(m) {{
    const baseDate = DATA{index}.baseDate || '';
    const dayOffset = Math.floor(m / 1440);
    const timeStr = formatTime{index}(m);
    
    if (dayOffset > 0 && baseDate) {{
      const parts = baseDate.split('.');
      if (parts.length === 3) {{
        const d = new Date(parts[2], parts[1] - 1, parseInt(parts[0]) + dayOffset);
        const newDate = d.getDate().toString().padStart(2, '0') + '.' + 
                        (d.getMonth() + 1).toString().padStart(2, '0') + '.' + 
                        d.getFullYear();
        return newDate + ' ' + timeStr;
      }}
    }}
    return baseDate + ' ' + timeStr;
  }}

  function renderTimeAxis{index}() {{
    const axis = document.getElementById('timeAxis{index}');
    if (!axis) return;
    axis.innerHTML = '';
    for (let m = 0; m <= TOTAL_MINUTES{index}; m += 15) {{
      const label = document.createElement('div');
      label.className = 'time-label';
      label.style.width = (15 * CONFIG{index}.MINUTE_WIDTH) + 'px';
      label.textContent = formatTime{index}(CONFIG{index}.WINDOW_START + m);
      axis.appendChild(label);
    }}
  }}

  window.scrollTimeline{index} = function(minutes) {{
    calcViewportMinutes{index}();
    const maxOffset = Math.max(0, TOTAL_MINUTES{index} - viewportMinutes{index});
    offset{index} = Math.max(0, Math.min(offset{index} + minutes, maxOffset));
    updateOffset{index}();
  }};

  function updateOffset{index}() {{
    calcViewportMinutes{index}();
    const px = offset{index} * CONFIG{index}.MINUTE_WIDTH;
    const container = document.getElementById('canvasContainer{index}');
    const axis = document.getElementById('timeAxis{index}');
    const btnLeft = document.getElementById('btnLeft{index}');
    const btnRight = document.getElementById('btnRight{index}');
    const timeRange = document.getElementById('timeRange{index}');
    
    if (container) container.style.transform = 'translateX(-' + px + 'px)';
    if (axis) axis.style.transform = 'translateX(-' + px + 'px)';
    
    const maxOffset = Math.max(0, TOTAL_MINUTES{index} - viewportMinutes{index});
    if (btnLeft) btnLeft.disabled = (offset{index} <= 0);
    if (btnRight) btnRight.disabled = (offset{index} >= maxOffset);
    
    const startMins = CONFIG{index}.WINDOW_START + offset{index};
    const endMins = startMins + 60;
    if (timeRange) timeRange.textContent = formatDateTime{index}(startMins) + ' — ' + formatTime{index}(endMins);
  }}

  window.toggleDetails{index} = function() {{
    const content = document.getElementById('detailsContent{index}');
    const icon = document.getElementById('detailsIcon{index}');
    if (!content || !icon) return;
    if (content.style.display === 'none' || content.style.display === '') {{
      content.style.display = 'block';
      icon.classList.add('open');
    }} else {{
      content.style.display = 'none';
      icon.classList.remove('open');
    }}
  }};

  window.toggleZone{index} = function(header) {{
    const content = header.nextElementSibling;
    const toggle = header.querySelector('.zone-toggle');
    if (!content || !toggle) return;
    if (content.style.display === 'none' || content.style.display === '') {{
      content.style.display = 'block';
      toggle.classList.add('open');
    }} else {{
      content.style.display = 'none';
      toggle.classList.remove('open');
    }}
  }};

  function showTooltip{index}(e) {{
    const tooltip = document.getElementById('tooltip{index}');
    if (!tooltip) return;
    const rect = e.target;
    const minute = parseInt(rect.dataset.minute);
    const timeStr = formatDateTime{index}(minute);
    const zone = parseInt(rect.dataset.zone);
    const tag = rect.dataset.tag;
    const desc = rect.dataset.desc || '';
    tooltip.innerHTML = '<div class="tooltip-header">' + timeStr + '</div>' +
      '<div class="tooltip-row"><span class="tooltip-label">Зона:</span><span class="tooltip-value">' + (DATA{index}.zoneNames[zone] || 'Зона ' + zone) + '</span></div>' +
      '<div class="tooltip-row"><span class="tooltip-label">Метка:</span><span class="tooltip-value">#' + tag + '</span></div>' +
      '<div class="tooltip-row"><span class="tooltip-label">Описание:</span><span class="tooltip-value">' + desc + '</span></div>';
    tooltip.classList.add('visible');
    moveTooltip{index}(e);
  }}

  function moveTooltip{index}(e) {{
    const tooltip = document.getElementById('tooltip{index}');
    if (!tooltip) return;
    const tooltipWidth = tooltip.offsetWidth || 280;
    const tooltipHeight = tooltip.offsetHeight || 150;
    const windowWidth = window.innerWidth;
    const windowHeight = window.innerHeight;
    
    let left = e.clientX + 15;
    let top = e.clientY + 15;
    
    // Проверяем правую границу
    if (left + tooltipWidth > windowWidth - 10) {{
      left = e.clientX - tooltipWidth - 15;
    }}
    // Проверяем нижнюю границу
    if (top + tooltipHeight > windowHeight - 10) {{
      top = e.clientY - tooltipHeight - 15;
    }}
    // Не уходим за левый край
    if (left < 10) left = 10;
    // Не уходим за верхний край
    if (top < 10) top = 10;
    
    tooltip.style.left = left + 'px';
    tooltip.style.top = top + 'px';
  }}

  function hideTooltip{index}() {{
    const tooltip = document.getElementById('tooltip{index}');
    if (tooltip) tooltip.classList.remove('visible');
  }}

  document.addEventListener('DOMContentLoaded', function() {{
    calcViewportMinutes{index}();
    renderTimeAxis{index}();
    updateOffset{index}();
    const svg = document.getElementById('timelineSvg{index}');
    if (svg) {{
      svg.querySelectorAll('.segment').forEach(function(el) {{
        el.addEventListener('mouseenter', showTooltip{index});
        el.addEventListener('mouseleave', hideTooltip{index});
        el.addEventListener('mousemove', moveTooltip{index});
      }});
    }}
  }});
  
  // Пересчитываем при изменении размера окна
  window.addEventListener('resize', function() {{
    calcViewportMinutes{index}();
    updateOffset{index}();
  }});
}})();
</script>
'''


def export_overall_excel_period(overall_by_date: dict):
    """Экспортирует данные в Excel файл"""
    if not overall_by_date:
        print("⚠️ Нет данных для Excel"); return
    dates_sorted = sorted(overall_by_date.keys())
    min_d = pd.to_datetime(min(dates_sorted)).strftime('%d.%m.%Y')
    max_d = pd.to_datetime(max(dates_sorted)).strftime('%d.%m.%Y')
    filename = f"AA_BLE Таймлайны SVG {min_d.replace('.','-')}—{max_d.replace('.','-')}.xlsx"
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Border, Side
    thin = Side(style='thin', color='FFAAAAAA')
    with pd.ExcelWriter(filename, engine='openpyxl') as w:
        for the_date in dates_sorted:
            df_x = overall_by_date[the_date].copy()
            if df_x is None or df_x.empty: continue
            events = df_x[['employee','area','ble_tag','zone_name','start','end','duration_minutes','tn_number']].rename(columns={
                'employee':'Сотрудник','area':'Участок','ble_tag':'Метка',
                'zone_name':'Зона','start':'Начало','end':'Конец',
                'duration_minutes':'Минуты','tn_number':'ТН'
            }).sort_values(['Сотрудник','Начало'])
            zone_pivot = df_x.pivot_table(index='employee', columns='zone_name', values='duration_minutes', aggfunc='sum', fill_value=0.0)
            zone_pivot = zone_pivot.applymap(round_051).reset_index().rename(columns={'employee':'Сотрудник'})
            tag = pd.to_datetime(the_date).strftime('%d.%m')
            sh_e = f"{tag} — Таблица"; sh_z = f"{tag} — Зоны"
            events.to_excel(w, index=False, sheet_name=sh_e)
            zone_pivot.to_excel(w, index=False, sheet_name=sh_z)
            for ws in (w.sheets[sh_e], w.sheets[sh_z]):
                ws.freeze_panes = 'A2'
                max_row = ws.max_row; max_col = ws.max_column
                for c in range(1, max_col + 1):
                    ws.column_dimensions[get_column_letter(c)].width = 20
                for r in range(1, max_row + 1):
                    for c in range(1, max_col + 1):
                        cell = ws.cell(row=r, column=c)
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    files.download(filename)
    print(f"✅ Excel '{filename}' скачан.")

def create_aable_presence_analysis_v2():
    """Главная функция для генерации SVG-таймлайнов (24ч, динамический viewport)"""
    print("=" * 80)
    print("🏢 ТАЙМЛАЙНЫ AA_BLE v2 (00:00–00:00, поминутно, SVG-диаграммы)")
    print("=" * 80)
    
    print("\n📈 ШАГ 1: Загрузка aa_ble…")
    per_ans = input("Период (например, 10.09.2025-14.09.2025, Enter — все): ").strip()
    date_from = date_to = None
    if per_ans:
        m = re.findall(r'(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})', per_ans)
        if len(m) >= 2: date_from, date_to = m[0], m[1]
    aable = load_aable_second_sheet_multi(date_from=date_from, date_to=date_to)
    if aable is None or aable.empty:
        print("❌ Нет данных aa_ble"); return
    
    print("\n📍 ШАГ 2: Журнал BLE (названия меток)…")
    tag_desc_map = attach_ble_desc_map(load_ble_journal())
    
    print("\n🗂 ШАГ 3: 'Привязка людей'…")
    people_mapping_df = load_people_mapping()
    if people_mapping_df is None:
        print("❌ Нельзя продолжить без 'Привязка людей'"); return
    area_map = build_area_map(people_mapping_df)
    fio_map = build_fio_map(people_mapping_df)
    
    aable['ble_tag'] = pd.to_numeric(aable['ble_tag'], errors='coerce')
    aable['zone_id'] = pd.to_numeric(aable['zone_id'], errors='coerce').fillna(0).astype(int)
    
    print("\n🔍 ШАГ 4: Поминутные отрезки по меткам…")
    segments = build_segments_all(aable, area_map, fio_map)
    if segments.empty:
        print("❌ Нет сегментов для построения"); return
    
    segments['area_key'] = segments['area'].fillna('').astype(str).str.strip().str.lower().str.replace(' ','', regex=False)
    all_dates = sorted(segments['date'].dropna().unique())
    clear_saved_charts()
    
    all_timeline_data = []
    overall_by_date = {}
    
    for the_date in all_dates:
        df_day = segments[segments['date'] == the_date].copy()
        if df_day.empty: continue
        
        file_date_mode = df_day['file_date'].dropna().astype(str)
        display_date = pd.to_datetime(file_date_mode.mode().iloc[0]).date() if not file_date_mode.empty else pd.to_datetime(the_date).date()
        date_str = format_file_date_for_title(display_date)
        print(f"\n🎨 Дата (из файла): {date_str}")
        
        # Без фильтрации по времени — используем все данные
        for emp, g in df_day.groupby('employee'):
            data = prepare_timeline_data(g, tag_desc_map)
            if data:
                all_timeline_data.append(data)
        
        overall_by_date[the_date] = df_day.copy()
    
    # Экспорт HTML
    if all_dates:
        if len(all_dates) == 1:
            date_str = format_file_date_for_title(pd.to_datetime(all_dates[0]).date())
            page_title = f"AA_BLE Таймлайны SVG — {date_str}"
            filename = f"AA_BLE Таймлайны SVG {date_str}.html"
        else:
            min_d = format_file_date_for_title(pd.to_datetime(min(all_dates)).date())
            max_d = format_file_date_for_title(pd.to_datetime(max(all_dates)).date())
            page_title = f"AA_BLE Таймлайны SVG — период {min_d}–{max_d}"
            filename = f"AA_BLE Таймлайны SVG {min_d}—{max_d}.html"
        
        export_combined_svg_html(all_timeline_data, tag_desc_map, filename, page_title)
    
    # Экспорт Excel
    export_overall_excel_period(overall_by_date)
    
    print("\n✅ ГОТОВО! SVG-таймлайны с динамическим viewport и KPI-карточками.")

# Стартовые сообщения
print("=" * 100)
print("🏢 AA_BLE ТАЙМЛАЙНЫ v2 (00:00–00:00, поминутно, SVG-диаграммы)")
print("=" * 100)
print("📊 Новые возможности:")
print("   • Таймлайн 00:00–00:00 (24ч) с поддержкой ночных смен")
print("   • Динамический viewport: от первой до последней минуты для каждого ТН")
print("   • time-display с датой: 27.12.2025 07:00 — 08:00")
print("   • KPI только за фактическое время присутствия")
print("   • SVG-диаграммы с навигацией < > (шаг 1 час)")
print("   • Полностью оффлайн (без CDN)")
print("=" * 100)
print()
print("✅ ГОТОВО! Запустите create_aable_presence_analysis_v2() для начала.")
