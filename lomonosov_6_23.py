# 🚀 ТАЙМЛАЙНЫ AA_BLE (окно 6:00–23:00, ПОМИНУТНО)
# - Диаграмма по меткам (каждая строка = 1 мин)
# - Цвет метки = оттенок по зоне:
#   1 зелёные, 2 оранжевые, 4 красные, 5 фиолетовые, 7 жёлтые, 10 синие, 0 серые, прочее — серые
# - Справа отдельная «карточка» со временем по ЗОНАМ за окно 6–23
print("🔄 Запуск анализа aa_ble (6:00–23:00, поминутно, панель времени по зонам)…")

# Для Google Colab
!pip install -q pandas plotly openpyxl odfpy

import io, re, warnings
from datetime import datetime, timedelta, time as dtime
import pandas as pd
import plotly.graph_objects as go
from google.colab import files
warnings.filterwarnings('ignore')

REEXPORT_CODE = "2025"
LAST_EXPORTS = []

def set_reexport_code(new_code):
    global REEXPORT_CODE
    REEXPORT_CODE = str(new_code)
    print("🔐 Код для повторной выгрузки обновлён.")

def list_saved_charts():
    if not LAST_EXPORTS:
        print("⚠️ Кэш пуст."); return
    print(f"📦 В кэше {len(LAST_EXPORTS)} диаграмм:")
    for i, it in enumerate(LAST_EXPORTS, 1):
        print(f"  {i}) дата={it['date_str']}, раздел={it['area_suffix']}, сотрудник={it.get('employee','')}")

def clear_saved_charts():
    LAST_EXPORTS.clear()
    print("🧹 Кэш диаграмм очищен.")

def redownload_last_charts(code=None):
    if code is None:
        code = input("Введите код для повторной выгрузки: ").strip()
    if str(code) != str(REEXPORT_CODE):
        print("❌ Неверный код."); return
    if not LAST_EXPORTS:
        print("⚠️ Нет сохранённых диаграмм."); return
    print("♻️ Повторная выгрузка диаграмм...")
    for it in LAST_EXPORTS:
        export_chart_auto(it['fig'], it['date_str'], area_suffix=it['area_suffix'], employee=it.get('employee'))
    print("✅ Повторная выгрузка завершена!")

def round_051(x):
    if pd.isna(x): return x
    try: val = float(x)
    except: return x
    frac = abs(val) - abs(int(val))
    up = abs(int(val)) + 1; dn = abs(int(val))
    res = up if frac >= 0.51 else dn
    return res if val >= 0 else -res

def parse_date(s):
    if pd.isna(s): return None
    # Если pandas уже прочитал как datetime — сразу возвращаем .date()
    if isinstance(s, (pd.Timestamp, datetime)):
        return s.date()
    # Иначе парсим: сначала пробуем yearfirst для ISO (2025-10-01), потом dayfirst
    d = pd.to_datetime(s, yearfirst=True, errors='coerce')
    if pd.isna(d):
        d = pd.to_datetime(s, dayfirst=True, errors='coerce')
    return d.date() if pd.notna(d) else None

def parse_time_only(val):
    if pd.isna(val): return None
    s = str(val).strip()
    for fmt in ('%H:%M:%S','%H:%M'):
        try: return datetime.strptime(s, fmt).time()
        except: pass
    try:
        v = float(s); base = datetime(1899,12,30)
        return (base + timedelta(days=v)).time()
    except: return None

def parse_date_from_filename(filename: str):
    name = str(filename)
    m = re.search(r'(\d{4})[ _.-](\d{2})[ _.-](\d{2})', name)
    if m:
        try: return datetime(int(m[1]), int(m[2]), int(m[3])).date()
        except: pass
    m = re.search(r'(\d{2})[ _.-](\d{2})[ _.-](\d{4})', name)
    if m:
        try: return datetime(int(m[3]), int(m[2]), int(m[1])).date()
        except: pass
    return None

def infer_file_date_from_df(df):
    """Определяет дату отчёта по ПЕРВОМУ информативному значению даты/времени в файле.
    Приоритет колонок: 'date'/'дата' → 'дата на объекте' → 'shift day'/'shift_day'/'день смены' → любой datetime-подобный из первой строки.
    Возвращает объект date или None."""
    try:
        cols = [str(c).strip() for c in df.columns]
        low = {i: str(c).strip().lower() for i, c in enumerate(cols)}
        def find_col(candidates):
            for i, name in low.items():
                for cand in candidates:
                    if name == cand or cand in name:
                        return i
            return None
        # 1) Ищем явную колонку даты (приоритет: date/дата, затем дата на объекте, затем день смены)
        ix_date = find_col(['date','дата'])
        ix_date_obj = find_col(['дата на объекте','date on site'])
        ix_shift = find_col(['shift day','shift_day','день смены','день'])

        print(f"🔍 Поиск даты: ix_date={ix_date}, ix_date_obj={ix_date_obj}, ix_shift={ix_shift}")

        for ix in [ix_date, ix_date_obj, ix_shift]:
            if ix is not None and ix < df.shape[1]:
                print(f"🔍 Пробуем колонку {ix} ({cols[ix]}), первые 3 значения: {df.iloc[:3, ix].tolist()}")
                s = pd.to_datetime(df.iloc[:, ix], errors='coerce')
                val = s.dropna().iloc[0] if not s.dropna().empty else None
                if pd.notna(val):
                    print(f"✅ Определена дата файла из колонки '{cols[ix]}': {val.date()}")
                    return val.date()
        # 2) Пробуем первую строку: берём первое значение, похожее на дату/время
        print("🔍 Пробуем найти дату в первой строке данных...")
        if df.shape[0] > 0:
            first_row = df.iloc[0, :]
            for i, v in enumerate(first_row.values):
                dt = pd.to_datetime(v, errors='coerce')
                if pd.notna(dt):
                    print(f"✅ Определена дата файла из первой строки, колонка {i} ({cols[i] if i < len(cols) else 'без имени'}): {dt.date()}")
                    return dt.date()
        print("⚠️ Не удалось определить дату файла из данных")
        return None
    except Exception as e:
        print(f"⚠️ Ошибка при определении даты файла: {e}")
        return None

def format_file_date_for_title(d):
    if not d: return ""
    return datetime.strftime(pd.to_datetime(d), "%d-%m-%Y")

def load_aable_second_sheet_multi(date_from=None, date_to=None):
    print("📈 Загрузите один или несколько файлов aa_ble (используется 2-й лист):")
    uploaded = files.upload()
    if not uploaded: return None
    frames = []
    for filename, raw in uploaded.items():
        try:
            if filename.endswith(('.xlsx','.xls')):
                xls = pd.ExcelFile(io.BytesIO(raw))
                sheet = xls.sheet_names[1] if len(xls.sheet_names) >= 2 else xls.sheet_names[0]
                df = xls.parse(sheet)
            elif filename.endswith('.csv'):
                df = pd.read_csv(io.StringIO(raw.decode('utf-8')))
            else:
                print(f"❌ Пропущен '{filename}' — неподдерживаемый формат"); continue
            df['__source__'] = filename
            inferred = infer_file_date_from_df(df)
            file_date = inferred or parse_date_from_filename(filename)
            df['__file_date__'] = file_date
            frames.append(df)
        except Exception as e:
            print(f"❌ Ошибка '{filename}': {e}")
    if not frames: return None
    all_df_raw = pd.concat(frames, ignore_index=True)
    print("🔎 Исходные колонки:", list(all_df_raw.columns))
    all_df = normalize_aable_columns(all_df_raw)
    all_df['shift_day'] = all_df['shift_day'].apply(parse_date)
    all_df['time_only'] = all_df['time_only'].apply(parse_time_only)

    # ОТЛАДКА: показываем, что получилось после парсинга
    print(f"🔍 После парсинга: записей={len(all_df)}")
    print(f"🔍 shift_day (уникальные значения): {all_df['shift_day'].dropna().unique()}")
    if '__file_date__' in all_df.columns:
        print(f"🔍 __file_date__ (определенная дата файла): {all_df['__file_date__'].dropna().unique()}")

    if date_from or date_to:
        dfrom = pd.to_datetime(date_from, dayfirst=True, errors='coerce').date() if date_from else None
        dto = pd.to_datetime(date_to, dayfirst=True, errors='coerce').date() if date_to else None
        mask = pd.Series(True, index=all_df.index)
        if dfrom: mask &= (all_df['shift_day'] >= dfrom)
        if dto:   mask &= (all_df['shift_day'] <= dto)
        all_df = all_df[mask]
        print(f"🔍 После фильтра по периоду: записей={len(all_df)}")

    # ВАЖНО: используем только записи того дня, который определён как дата файла (первая информативная дата)
    if '__file_date__' in all_df.columns:
        fd_series = pd.to_datetime(all_df['__file_date__'], errors='coerce').dt.date
        print(f"🔍 Фильтруем: оставляем только записи где shift_day == {fd_series.unique() if hasattr(fd_series, 'unique') else fd_series}")
        all_df = all_df[all_df['shift_day'] == fd_series]
        print(f"🔍 После фильтра по дате файла: записей={len(all_df)}")

    print(f"✅ aa_ble: записей={len(all_df)}")
    print(all_df.head(3))
    return all_df

def load_ble_journal():
    print("📖 Загрузите 'Журнал BLE' (метка → описание/локация):")
    uploaded = files.upload()
    if not uploaded: return None
    filename = list(uploaded.keys())[0]
    try:
        if filename.endswith(('.xlsx','.xls')):
            df = pd.read_excel(io.BytesIO(uploaded[filename]))
        elif filename.endswith('.csv'):
            df = pd.read_csv(io.StringIO(uploaded[filename].decode('utf-8')))
        else:
            print("❌ Неподдерживаемый формат журнала BLE"); return None
        print(f"✅ Журнал BLE загружен: {len(df)} записей")
        return df
    except Exception as e:
        print(f"❌ Ошибка чтения журнала BLE: {e}")
        return None

def load_people_mapping():
    print("🗂 Загрузите 'Привязка людей' (A: ТН, B: ФИО, D: Участок Работ):")
    uploaded = files.upload()
    if not uploaded: return None
    filename = list(uploaded.keys())[0]
    try:
        if filename.endswith(('.xlsx','.xls')):
            df = pd.read_excel(io.BytesIO(uploaded[filename]))
        elif filename.endswith('.csv'):
            df = pd.read_csv(io.StringIO(uploaded[filename].decode('utf-8')))
        else:
            print("❌ Неподдерживаемый формат файла привязки"); return None
        print(f"✅ Привязка людей загружена: {len(df)} записей")
        return df
    except Exception as e:
        print(f"❌ Ошибка чтения файла привязки: {e}")
        return None

def build_area_map(mapping_df):
    area_map = {}
    if mapping_df is None or mapping_df.empty: return area_map
    for _, row in mapping_df.iterrows():
        try:
            tn = str(row.iloc[0]).strip()
            area = str(row.iloc[3]).strip() if len(row) > 3 else ""
            if tn and area and area.lower() != 'nan':
                area_map[tn] = area
        except: continue
    print(f"✅ Сформирована карта участков: {len(area_map)} сотрудников")
    return area_map

def build_fio_map(mapping_df):
    """Создает карту ТН -> ФИО из колонки B файла 'Привязка людей'"""
    fio_map = {}
    if mapping_df is None or mapping_df.empty: return fio_map
    for _, row in mapping_df.iterrows():
        try:
            tn = str(row.iloc[0]).strip()
            fio = str(row.iloc[1]).strip() if len(row) > 1 else ""  # Колонка B (индекс 1)
            if tn and fio and fio.lower() != 'nan':
                fio_map[tn] = fio
        except: continue
    print(f"✅ Сформирована карта ФИО: {len(fio_map)} сотрудников")
    return fio_map

def normalize_aable_columns(df):
    orig = df.copy()
    cols = list(orig.columns)
    low = {i: str(c).strip().lower() for i, c in enumerate(cols)}
    def find_by_name(needles):
        for i, name in low.items():
            for n in needles:
                if name == n: return i
        for i, name in low.items():
            for n in needles:
                if n in name: return i
        return None
    POS = {'tn_number':0,'shift_day':3,'ble_tag':10,'zone_id':11,'time_only':15}
    IX_TN   = find_by_name(['тн','tn','таб','табель','tabnum','tab_num','personal'])
    IX_DAY  = find_by_name(['день смены','shift day','shift_day','день','смены'])
    IX_TAG  = find_by_name(['ble','метк','mekta','metka','маяч','tag'])
    IX_ZONE = find_by_name(['zona','зона','zone'])
    IX_TIME = find_by_name(['время на объекте','время','time'])
    def pick(ix_name, pos_key):
        ix = {'tn_number':IX_TN,'shift_day':IX_DAY,'ble_tag':IX_TAG,'zone_id':IX_ZONE,'time_only':IX_TIME}[ix_name]
        if ix is not None and ix < orig.shape[1]: return orig.iloc[:, ix]
        pos = POS[pos_key]
        return orig.iloc[:, pos] if pos < orig.shape[1] else pd.Series([None]*len(orig))
    out = pd.DataFrame({
        'tn_number': pick('tn_number','tn_number'),
        'shift_day': pick('shift_day','shift_day'),
        'ble_tag': pick('ble_tag','ble_tag'),
        'zone_id': pick('zone_id','zone_id'),
        'time_only': pick('time_only','time_only'),
    })
    if '__source__' in orig.columns: out['__source__'] = orig['__source__']
    if '__file_date__' in orig.columns: out['__file_date__'] = orig['__file_date__']
    def src_name(s):
        try: return cols[s] if s is not None and s < len(cols) else None
        except: return None
    print("🧭 Маппинг колонок:",
          f"\n  tn_number ← {src_name(IX_TN) or 'A'}",
          f"\n  shift_day ← {src_name(IX_DAY) or 'D'}",
          f"\n  ble_tag   ← {src_name(IX_TAG) or 'K'}",
          f"\n  zone_id   ← {src_name(IX_ZONE) or 'L'}",
          f"\n  time_only ← {src_name(IX_TIME) or 'P'}")
    return out

def attach_ble_desc_map(ble_journal_df):
    if ble_journal_df is None or ble_journal_df.empty: return {}
    jr = ble_journal_df.copy()
    jr.iloc[:,0] = jr.iloc[:,0].astype(str).str.strip()
    if jr.shape[1] < 4: return {}
    jr['__desc__'] = jr.iloc[:,3].astype(str).str.strip()
    d = {}
    for _, r in jr.iterrows():
        key = r.iloc[0]; val = r['__desc__']
        if str(val).strip(): d[str(key)] = val
    return d

PALETTES = {
    1: ['#27AE60','#2ECC71','#229954','#1E8449','#58D68D'],  # Зоны проведения работ - зеленый
    4: ['#E74C3C','#C0392B','#E53935','#D32F2F','#FF6F61'],  # Курилки - красный
    10:['#1F77B4','#2980B9','#3498DB','#2874A6','#5DADE2'],  # Зона выдачи WW - синий
    13:['#E67E22','#F39C12','#D35400','#E67E22','#F1C40F'],  # КПП - оранжевый
    2: ['#E67E22','#F39C12','#D35400','#E67E22','#F1C40F'],
    5: ['#8E44AD','#9B59B6','#7D3C98','#6C3483','#AF7AC5'],
    7: ['#F1C40F','#F4D03F','#F7DC6F','#D4AC0D','#F9E79F'],
    0: ['#95A5A6','#7F8C8D','#B0B0B0','#A0A0A0','#BDC3C7'],
    'other': ['#95A5A6','#7F8C8D','#B0B0B0','#A0A0A0','#BDC3C7']
}
def color_for_tag_zone(tag, zone_id):
    try: zid = int(zone_id) if pd.notna(zone_id) else 0
    except: zid = 0
    pal = PALETTES.get(zid, PALETTES['other'])
    try: idx = abs(int(float(tag))) % len(pal)
    except: idx = 0
    return pal[idx]

# ROW→MINUTE: каждая строка = 1 минута для метки
def build_user_minute_bins_by_tag(df_user):
    rows = df_user.dropna(subset=['shift_day','time_only']).sort_values(['shift_day','time_only']).to_dict('records')
    if not rows: return []
    minute_bins = []; prev_dt = None; current_day = rows[0]['shift_day']
    for r in rows:
        day = r['shift_day'] or current_day
        t = r['time_only']
        if t is None: continue
        dt = datetime.combine(day, t)
        if prev_dt and dt < prev_dt:
            dt = datetime.combine(day + timedelta(days=1), t)
            current_day = day + timedelta(days=1)
        else:
            current_day = day
        try: tag_val = int(float(str(r.get('ble_tag')))) if r.get('ble_tag') not in [None,""] else 0
        except: tag_val = 0
        try: zone_val = int(r.get('zone_id')) if pd.notna(r.get('zone_id')) else 0
        except: zone_val = 0
        minute_bins.append({'start': dt, 'end': dt + timedelta(minutes=1), 'ble_tag': tag_val, 'zone_id': zone_val})
        prev_dt = dt
    return minute_bins

def build_segments_all(df, area_map, fio_map):
    out_rows = []
    for (tn, shift_day), g in df.groupby(['tn_number','shift_day']):
        bins = build_user_minute_bins_by_tag(g)
        file_dates = g['__file_date__'].dropna().astype(str)
        file_date = file_dates.mode().iloc[0] if not file_dates.empty else None
        for b in bins:
            area = area_map.get(str(tn).strip(), '')
            fio = fio_map.get(str(tn).strip(), f"ТН {str(tn).strip()}")  # Fallback к ТН если ФИО не найдено
            zid = b['zone_id'] if pd.notna(b['zone_id']) else 0
            out_rows.append({
                'tn_number': str(tn).strip(),
                'employee': fio,
                'area': area,
                'date': b['start'].date(),
                'file_date': file_date,
                'start': b['start'],
                'end': b['end'],
                'duration_minutes': 1.0,
                'ble_tag': b['ble_tag'],
                'zone_id': zid,
                'zone_name': {0:"Вне зоны BLE-маячков",1:"Зоны проведения работ",2:"Столовые",3:"Опасные зоны",4:"Курилки",5:"Зоны отдыха",6:"ВЖГ",7:"Туалеты",8:"Остановки автобусов",9:"Административные помещения",10:"Зона выдачи WW",11:"Склад",12:"Мастерские",13:"КПП"}.get(zid, f"Зона {zid}"),
            })
    out = pd.DataFrame(out_rows)
    if not out.empty:
        out = out.sort_values(['date','employee','start'])
        out['duration_minutes'] = out['duration_minutes'].apply(round_051)
    return out

def report_zero_metka(df_raw):
    s = df_raw.get('ble_tag')
    if s is None:
        print("⚠️ Колонка метки BLE отсутствует."); return
    mask0 = s.isna() | (s.astype(str).str.strip().isin(['0','0.0','']))
    counts = df_raw.loc[mask0].groupby('tn_number').size().sort_values(ascending=False)
    offenders = counts[counts > 100]
    if offenders.empty:
        print("✅ Нет сотрудников с >100 строк метки 0.")
    else:
        print("❗ >100 строк метки 0 (Вне зоны BLE):")
        print(offenders.to_string())
    return offenders

def log_time_gaps(df_raw, fio_map):
    """Логирует разрывы во времени для каждого сотрудника (ТН) по сменам"""
    if df_raw is None or df_raw.empty:
        print("⚠️ Нет данных для анализа разрывов")
        return

    print("\n" + "=" * 70)
    print("🔍 АНАЛИЗ РАЗРЫВОВ ВО ВРЕМЕНИ ПО СОТРУДНИКАМ")
    print("=" * 70)

    # Группируем по ТН и дню смены
    gaps_found = False
    for tn, tn_group in df_raw.groupby('tn_number'):
        tn_str = str(tn).strip()
        fio = fio_map.get(tn_str, f"ТН {tn_str}")

        # Статистика по сменам для этого ТН
        shift_stats = {}
        shift_gaps = {}

        for shift_day, day_group in tn_group.groupby('shift_day'):
            if pd.isna(shift_day):
                continue

            # Получаем времена и сортируем
            times = day_group['time_only'].dropna().tolist()
            times_sorted = sorted([t for t in times if t is not None])

            if len(times_sorted) < 2:
                continue

            # Считаем количество минут в смене
            shift_stats[shift_day] = len(times_sorted)

            # Ищем разрывы
            gaps = []
            for i in range(1, len(times_sorted)):
                prev_time = times_sorted[i-1]
                curr_time = times_sorted[i]

                # Конвертируем в минуты от полуночи для сравнения
                prev_mins = prev_time.hour * 60 + prev_time.minute
                curr_mins = curr_time.hour * 60 + curr_time.minute

                diff = curr_mins - prev_mins
                if diff > 1:  # Разрыв больше 1 минуты
                    expected_time = (datetime.combine(datetime.today(), prev_time) + timedelta(minutes=1)).time()
                    gaps.append({
                        'from': prev_time.strftime('%H:%M'),
                        'to': curr_time.strftime('%H:%M'),
                        'expected': expected_time.strftime('%H:%M'),
                        'gap_minutes': diff - 1
                    })

            if gaps:
                shift_gaps[shift_day] = gaps

        # Если есть разрывы для этого ТН — выводим
        if shift_gaps:
            gaps_found = True
            # Находим самую большую смену
            if shift_stats:
                max_shift = max(shift_stats, key=shift_stats.get)
                max_shift_mins = shift_stats[max_shift]
                max_shift_str = max_shift.strftime('%d.%m.%Y') if hasattr(max_shift, 'strftime') else str(max_shift)
            else:
                max_shift_str = "н/д"
                max_shift_mins = 0

            print(f"\n👤 {fio} (ТН: {tn_str})")
            print(f"   📊 Самая большая смена: {max_shift_str} ({max_shift_mins} мин)")

            for shift_day, gaps in shift_gaps.items():
                shift_str = shift_day.strftime('%d.%m.%Y') if hasattr(shift_day, 'strftime') else str(shift_day)
                shift_mins = shift_stats.get(shift_day, 0)
                print(f"   📅 Смена {shift_str} ({shift_mins} мин):")
                for gap in gaps:
                    print(f"      ⚠️ Разрыв: {gap['from']} → {gap['to']} (ожидалось {gap['expected']}, пропущено {gap['gap_minutes']} мин)")

    if not gaps_found:
        print("✅ Разрывов во времени не обнаружено.")

    print("=" * 70 + "\n")

def analyze_tag_zones(df_raw):
    """Анализирует распределение меток по зонам"""
    if df_raw is None or df_raw.empty:
        print("⚠️ Нет данных для анализа зон меток")
        return

    print("\n🔍 АНАЛИЗ МЕТОК ПО ЗОНАМ:")
    print("=" * 50)

    # Группируем по метке и зоне
    tag_zone_analysis = df_raw.groupby(['ble_tag', 'zone_id']).size().reset_index(name='count')
    tag_zone_analysis = tag_zone_analysis.sort_values(['ble_tag', 'count'], ascending=[True, False])

    # Показываем топ-10 меток с их зонами
    top_tags = tag_zone_analysis['ble_tag'].value_counts().head(10)

    for tag in top_tags.index:
        if pd.notna(tag) and str(tag).strip() not in ['0', '0.0', '']:
            tag_data = tag_zone_analysis[tag_zone_analysis['ble_tag'] == tag]
            print(f"\n📌 Метка {tag}:")
            for _, row in tag_data.iterrows():
                zone_id = int(row['zone_id']) if pd.notna(row['zone_id']) else 0
                zone_name = {0:"Вне зоны BLE-маячков",1:"Зоны проведения работ",2:"Столовые",3:"Опасные зоны",4:"Курилки",5:"Зоны отдыха",6:"ВЖГ",7:"Туалеты",8:"Остановки автобусов",9:"Административные помещения",10:"Зона выдачи WW",11:"Склад",12:"Мастерские",13:"КПП"}.get(zone_id, f"Зона {zone_id}")
                print(f"   Зона {zone_id} ({zone_name}): {int(row['count'])} записей")

    # Специально для метки 1
    if 1 in df_raw['ble_tag'].values:
        print(f"\n🎯 ДЕТАЛЬНЫЙ АНАЛИЗ МЕТКИ 1:")
        tag1_data = df_raw[df_raw['ble_tag'] == 1]
        if not tag1_data.empty:
            zone_dist = tag1_data['zone_id'].value_counts().sort_index()
            for zone_id, count in zone_dist.items():
                zone_name = {0:"Вне зоны BLE-маячков",1:"Зоны проведения работ",2:"Столовые",3:"Опасные зоны",4:"Курилки",5:"Зоны отдыха",6:"ВЖГ",7:"Туалеты",8:"Остановки автобусов",9:"Административные помещения",10:"Зона выдачи WW",11:"Склад",12:"Мастерские",13:"КПП"}.get(zone_id, f"Зона {zone_id}")
                print(f"   Зона {zone_id} ({zone_name}): {count} записей")

            # Показываем временное распределение
            if 'time_only' in tag1_data.columns:
                print(f"\n⏰ Временное распределение метки 1:")
                tag1_data['hour'] = pd.to_datetime(tag1_data['time_only'], errors='coerce').dt.hour
                hourly_zones = tag1_data.groupby(['hour', 'zone_id']).size().reset_index(name='count')
                for hour in sorted(hourly_zones['hour'].dropna().unique()):
                    hour_data = hourly_zones[hourly_zones['hour'] == hour]
                    zones_str = ", ".join([f"зона{int(z)}" for z in hour_data['zone_id']])
                    print(f"   {int(hour):02d}:00 - {zones_str}")

    return tag_zone_analysis

def lighten_color(hex_color, factor=0.3):
    """Осветляет hex цвет на заданный фактор (0-1)"""
    hex_color = hex_color.lstrip('#')
    r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
    r = int(r + (255 - r) * factor)
    g = int(g + (255 - g) * factor)
    b = int(b + (255 - b) * factor)
    return f'#{r:02x}{g:02x}{b:02x}'

def create_user_timeline_chart_window(df_user, tag_desc_map, row_height=60, window_start=(6,0), window_end=(23,0)):
    if df_user is None or df_user.empty: return None, None
    df = df_user.copy()
    df['start'] = pd.to_datetime(df['start']); df['end'] = pd.to_datetime(df['end'])
    file_date = df.get('file_date')
    file_date = file_date.iloc[0] if file_date is not None and not file_date.empty else None
    display_date = pd.to_datetime(file_date).date() if file_date else df['start'].dt.date.min()
    ws = datetime.combine(display_date, dtime(window_start[0], window_start[1]))
    we = datetime.combine(display_date, dtime(window_end[0], window_end[1]))
    df = df[(df['start'] >= ws) & (df['start'] < we)].copy()
    if df.empty: return None, None
    df['duration_ms'] = (df['end'] - df['start']).dt.total_seconds() * 1000
    user = df['employee'].iloc[0]
    total_by_tag = df.groupby(['ble_tag','zone_id'])['duration_minutes'].sum().to_dict()
    zone_sum = df.groupby('zone_id')['duration_minutes'].sum().apply(round_051).sort_index()
    
    # Собираем данные по меткам внутри каждой зоны
    zone_tags_data = {}
    for (tag, zid), minutes in total_by_tag.items():
        if zid not in zone_tags_data:
            zone_tags_data[zid] = []
        tag_int = int(tag) if pd.notna(tag) else 0
        mins = int(round_051(minutes))
        zone_tags_data[zid].append({'tag': tag_int, 'minutes': mins})
    # Сортируем метки внутри каждой зоны по убыванию времени
    for zid in zone_tags_data:
        zone_tags_data[zid] = sorted(zone_tags_data[zid], key=lambda x: x['minutes'], reverse=True)

    fig = go.Figure()
    for (tag, z), g in df.groupby(['ble_tag','zone_id']):
        tag_int = int(tag) if pd.notna(tag) else 0
        tag_name = "Вне зоны BLE-маячков" if tag_int == 0 else f"Метка #{tag_int}"
        desc = tag_desc_map.get(str(tag_int), "").strip()
        title = f"{tag_name} — {desc}" if desc else tag_name
        color = color_for_tag_zone(tag_int, z)
        tot = int(round_051(total_by_tag.get((tag, z), 0)))
        hovers = []
        for _, r in g.iterrows():
            hovers.append("<br>".join([
                f"{r['employee']}",
                f"Участок: {r.get('area','')}" if pd.notna(r.get('area')) and str(r.get('area')).strip() else "",
                title,
                f"Зона: {r['zone_name']}",
                f"Всего в этой метке (6–23): {tot} мин",
                f"Начало: {r['start'].strftime('%d.%m %H:%M')}",
                f"Конец: {r['end'].strftime('%d.%m %H:%M')}",
                "Длительность: 1 мин"
            ]))
        fig.add_trace(go.Bar(
            y=[user]*len(g), x=g['duration_ms'], base=g['start'], orientation='h',
            marker=dict(color=color, line=dict(color='rgba(0,0,0,0.15)', width=0.6)),
            hovertemplate='%{customdata}<extra></extra>', customdata=hovers, showlegend=False
        ))

    date_str = format_file_date_for_title(display_date)
    area_label = df['area'].iloc[0] if str(df['area'].iloc[0]).strip() else 'Без участка'
    fig.update_layout(
        title=dict(text=f"{user} — {date_str}", x=0.5, font=dict(size=20, color='#333')),
        width=980, height=max(360, row_height + 230),
        plot_bgcolor='white',
        paper_bgcolor='rgba(0,0,0,0)',
        margin=dict(l=110, r=20, t=70, b=80),
        barmode='overlay', hoverlabel=dict(font=dict(size=10))
    )
    fig.update_xaxes(
        title=dict(text="Время (6:00–23:00)", font=dict(size=16, color="#333")),
        type='date', tickformat="%H:%M", dtick=3600000,
        range=[ws, we], showgrid=True, gridcolor="#e0e0e0", gridwidth=1,
        linecolor="#333", linewidth=2
    )
    fig.update_yaxes(
        title=dict(text="Сотрудник", font=dict(size=16, color="#333")),
        tickfont=dict(size=13, color="#333"), categoryorder="array", categoryarray=[user],
        showgrid=True, gridcolor="#e0e0e0", gridwidth=1, linecolor="#333", linewidth=2
    )

    # Панель справа: время по зонам (6–23) с accordion для меток
    zone_names = {0:"Вне зоны BLE-маячков",1:"Зоны проведения работ",2:"Столовые",3:"Опасные зоны",4:"Курилки",5:"Зоны отдыха",6:"ВЖГ",7:"Туалеты",8:"Остановки автобусов",9:"Административные помещения",10:"Зона выдачи WW",11:"Склад",12:"Мастерские",13:"КПП"}
    order = [1,4,10,13,2,7,0] + [z for z in zone_sum.index if z not in [1,4,10,13,2,7,0]]
    
    # Общее время присутствия сотрудника (сумма всех зафиксированных минут)
    total_presence = int(zone_sum.sum())
    
    items = []
    for zid in order:
        if zid not in zone_sum.index: continue
        mins = int(zone_sum.loc[zid])
        pct = round(mins / total_presence * 100, 1) if total_presence > 0 else 0
        name = zone_names.get(zid, f"Зона {zid}")
        color = PALETTES.get(zid, PALETTES['other'])[0]
        light_color = lighten_color(color, 0.4)
        
        # Формируем список меток для этой зоны
        tags_in_zone = zone_tags_data.get(zid, [])
        tags_html = ""
        if tags_in_zone:
            tag_items = []
            for t in tags_in_zone:
                tag_int = t['tag']
                tag_mins = t['minutes']
                tag_pct = round(tag_mins / total_presence * 100, 1) if total_presence > 0 else 0
                tag_name = "Вне зоны" if tag_int == 0 else f"#{tag_int}"
                desc = tag_desc_map.get(str(tag_int), "").strip()
                tag_label = f"{tag_name} ({desc})" if desc else tag_name
                tag_items.append(f"<div class='tag-row' style='background:{light_color}'><span class='tag-name'>{tag_label}</span><span class='tag-val'>{tag_mins} мин ({tag_pct}%)</span></div>")
            tags_html = "".join(tag_items)
        
        items.append(f"""
        <div class='zone-accordion'>
          <div class='zone-header' onclick='toggleZone(this)'>
            <span class='zone-toggle'>▶</span>
            <span class='dot' style='background:{color}'></span>
            <span class='zname'>{name}</span>
            <span class='zval'>{mins} мин ({pct}%)</span>
          </div>
          <div class='zone-content' style='display:none;'>
            {tags_html if tags_html else "<div class='tag-empty'>Нет меток</div>"}
          </div>
        </div>""")
    
    panel_html = f"""
    <div class="zone-card">
      <div class="zone-title">Время по зонам (6:00–23:00)</div>
      {''.join(items) if items else "<div class='zone-empty'>Нет данных</div>"}
    </div>
    """
    return fig, panel_html

def export_chart_auto(fig, date_str, area_suffix=None, employee=None):
    if not fig: return
    area_label = str(area_suffix).strip() if area_suffix else "Общая"
    user_label = f" — {employee}" if employee else ""
    safe_date = str(date_str).replace('/','-')
    filename = f"AA_BLE Таймлайн {area_label} {safe_date}{user_label}.html"
    try:
        page_title = f"AA_BLE Таймлайн — {area_label} — {date_str}{user_label}"
        html_string = fig.to_html(include_plotlyjs='cdn', config={
            'displayModeBar': True, 'displaylogo': False,
            'toImageButtonOptions': {'format':'png','filename':filename.replace('.html',''),
                                     'height': int(fig.layout.height or 800),
                                     'width': int(fig.layout.width or 980), 'scale': 2},
            'responsive': True
        })
        custom_html = html_string.replace('<head>', f'''<head>
<meta charset="utf-8"><title>{page_title}</title>
<style>body{{font-family:Arial;margin:0;padding:20px;background:#f8f9fa;}}
.plotly-graph-div{{background:#fff;border-radius:8px;box-shadow:0 2px 10px rgba(0,0,0,0.1);padding:20px;}}</style>''')
        with open(filename,'w',encoding='utf-8') as f: f.write(custom_html)
        files.download(filename)
        print(f"✅ HTML '{filename}' скачан")
    except Exception as e:
        print(f"❌ Ошибка экспорта: {e}")

def fig_to_div(fig, title_text):
    return fig.to_html(include_plotlyjs=False, full_html=False, config={
        'displayModeBar': True, 'displaylogo': False,
        'toImageButtonOptions': {'format':'png','filename':title_text.replace(' ','_'),
                                 'height': int(fig.layout.height or 800),
                                 'width': int(fig.layout.width or 980), 'scale': 2},
        'responsive': True
    })

def export_combined_html(sections, filename, page_title):
    toc_items, content = [], []
    for i, sec in enumerate(sections, 1):
        anchor = f"sec{i}"
        h = sec.get('heading') or ''
        sh = sec.get('subheading') or ''
        toc_items.append(f'<li><a href="#{anchor}">{h}{" — " + sh if sh else ""}</a></li>')
        content.append(f'''
<section id="{anchor}">
  <div class="back"><a href="#toc">← К оглавлению</a></div>
  <h2>{h}{(" — " + sh) if sh else ""}</h2>
  <div class="row">
    <div class="panel">{sec['panel_html']}</div>
    <div class="plot">{sec['fig_div']}</div>
  </div>
  <hr/>
</section>
''')
    html = f'''<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="utf-8"/>
<title>{page_title}</title>
<script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
<style>
 body {{ font-family: Arial, sans-serif; background:#f8f9fa; margin:0; padding:0; }}
 .container {{ max-width: 1320px; margin: 0 auto; padding: 24px; }}
 .card {{ background:#fff; border-radius:8px; box-shadow:0 2px 10px rgba(0,0,0,0.1); padding:20px; }}
 h1 {{ margin:0 0 16px; color:#333; }}
 h2 {{ margin:12px 0; color:#333; }}
 ul.toc > li {{ margin:6px 0; }}
 a {{ color:#0d6efd; text-decoration:none; }} a:hover {{ text-decoration:underline; }}
 .row {{ display:flex; gap:16px; align-items:flex-start; }}
 .panel {{ width:300px; flex:0 0 300px; }}
 .plot {{ flex:1; min-width:0; }}
 .zone-card {{ background:#fff; border-radius:8px; box-shadow:0 2px 10px rgba(0,0,0,0.1); padding:14px; }}
 .zone-title {{ font-weight:bold; margin:4px 0 10px; }}
 .zone-accordion {{ margin-bottom:2px; }}
 .zone-header {{ display:flex; align-items:center; padding:8px 6px; cursor:pointer; border-radius:4px; transition:background 0.2s; }}
 .zone-header:hover {{ background:#f0f0f0; }}
 .zone-toggle {{ font-size:10px; margin-right:6px; color:#666; transition:transform 0.2s; }}
 .zone-toggle.open {{ transform:rotate(90deg); }}
 .dot {{ display:inline-block; width:12px; height:12px; border-radius:50%; margin-right:8px; flex-shrink:0; }}
 .zname {{ flex:1; font-size:13px; }}
 .zval {{ margin-left:12px; font-weight:bold; color:#333; font-size:13px; }}
 .zone-content {{ margin-left:28px; padding-left:8px; border-left:2px solid #e0e0e0; }}
 .tag-row {{ display:flex; align-items:center; justify-content:space-between; padding:5px 8px; margin:2px 0; border-radius:4px; font-size:12px; }}
 .tag-name {{ flex:1; color:#444; }}
 .tag-val {{ margin-left:8px; font-weight:600; color:#333; }}
 .tag-empty {{ font-size:11px; color:#999; padding:4px 8px; }}
 .zone-empty {{ font-size:12px; color:#999; padding:8px; }}
 .back {{ margin: 8px 0 4px }}
 hr {{ margin: 24px 0; border:0; border-top:1px solid #ddd }}
</style>
</head>
<body>
<div class="container">
  <div class="card" id="toc">
    <h1>{page_title}</h1>
    <h3 style="margin-top:0">Оглавление</h3>
    <ul class="toc">{''.join(toc_items)}</ul>
  </div>
  <div style="height:16px"></div>
  {''.join(content)}
</div>
<script>
function toggleZone(header) {{
  var content = header.nextElementSibling;
  var toggle = header.querySelector('.zone-toggle');
  if (content.style.display === 'none') {{
    content.style.display = 'block';
    toggle.classList.add('open');
  }} else {{
    content.style.display = 'none';
    toggle.classList.remove('open');
  }}
}}
</script>
</body>
</html>'''
    with open(filename, 'w', encoding='utf-8') as f: f.write(html)
    files.download(filename)
    print(f"✅ Объединённый HTML '{filename}' скачан.")

def export_overall_excel_period(overall_by_date: dict):
    if not overall_by_date:
        print("⚠️ Нет данных для Excel"); return
    dates_sorted = sorted(overall_by_date.keys())
    min_d = pd.to_datetime(min(dates_sorted)).strftime('%d.%m.%Y')
    max_d = pd.to_datetime(max(dates_sorted)).strftime('%d.%m.%Y')
    filename = f"AA_BLE Таймлайны 6-23 {min_d.replace('.','-')}—{max_d.replace('.','-')}.xlsx"
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Border, Side
    thin = Side(style='thin', color='FFAAAAAA')
    with pd.ExcelWriter(filename, engine='openpyxl') as w:
        for the_date in dates_sorted:
            df_x = overall_by_date[the_date].copy()
            if df_x is None or df_x.empty: continue
            events = df_x[['employee','area','ble_tag','zone_name','start','end','duration_minutes','tn_number']].rename(columns={
                'employee':'Сотрудник','area':'Участок','ble_tag':'Метка',
                'zone_name':'Зона','start':'Начало','end':'Конец',
                'duration_minutes':'Минуты','tn_number':'ТН'
            }).sort_values(['Сотрудник','Начало'])
            zone_pivot = df_x.pivot_table(index='employee', columns='zone_name', values='duration_minutes', aggfunc='sum', fill_value=0.0)
            zone_pivot = zone_pivot.applymap(round_051).reset_index().rename(columns={'employee':'Сотрудник'})
            tag = pd.to_datetime(the_date).strftime('%d.%m')
            sh_e = f"{tag} — Таблица 6-23"; sh_z = f"{tag} — Зоны 6-23"
            events.to_excel(w, index=False, sheet_name=sh_e)
            zone_pivot.to_excel(w, index=False, sheet_name=sh_z)
            for ws in (w.sheets[sh_e], w.sheets[sh_z]):
                ws.freeze_panes = 'A2'
                max_row = ws.max_row; max_col = ws.max_column
                for c in range(1, max_col + 1):
                    ws.column_dimensions[get_column_letter(c)].width = 20
                for r in range(1, max_row + 1):
                    for c in range(1, max_col + 1):
                        cell = ws.cell(row=r, column=c)
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    files.download(filename)
    print(f"✅ Excel '{filename}' скачан.")

def create_aable_presence_analysis():
    print("=" * 80)
    print("🏢 ТАЙМЛАЙНЫ AA_BLE (6:00–23:00, поминутно; справа — окно «Время по зонам»)")
    print("=" * 80)
    try:
        row_height = int(input("Высота строки сотрудника (пикс), Enter=60: ").strip() or "60")
    except: row_height = 60
    print("\n📈 ШАГ 1: Загрузка aa_ble…")
    per_ans = input("Период (например, 10.09.2025-14.09.2025, Enter — все): ").strip()
    date_from = date_to = None
    if per_ans:
        m = re.findall(r'(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})', per_ans)
        if len(m) >= 2: date_from, date_to = m[0], m[1]
    aable = load_aable_second_sheet_multi(date_from=date_from, date_to=date_to)
    if aable is None or aable.empty:
        print("❌ Нет данных aa_ble"); return
    print("\n📍 ШАГ 2: Журнал BLE (названия меток)…")
    tag_desc_map = attach_ble_desc_map(load_ble_journal())
    print("\n🗂 ШАГ 3: 'Привязка людей'…")
    people_mapping_df = load_people_mapping()
    if people_mapping_df is None:
        print("❌ Нельзя продолжить без 'Привязка людей'"); return
    area_map = build_area_map(people_mapping_df)
    fio_map = build_fio_map(people_mapping_df)
    aable['ble_tag'] = pd.to_numeric(aable['ble_tag'], errors='coerce')
    aable['zone_id'] = pd.to_numeric(aable['zone_id'], errors='coerce').fillna(0).astype(int)
    print("\n🧭 Проверка 'метка 0'…"); _ = report_zero_metka(aable)
    print("\n🔍 Анализ разрывов во времени…"); log_time_gaps(aable, fio_map)
    print("\n🔍 Анализ меток по зонам…"); _ = analyze_tag_zones(aable)
    print("\n🔍 ШАГ 4: Поминутные отрезки по меткам…")
    segments = build_segments_all(aable, area_map, fio_map)
    if segments.empty:
        print("❌ Нет сегментов для построения"); return
    segments['area_key'] = segments['area'].fillna('').astype(str).str.strip().str.lower().str.replace(' ','', regex=False)
    all_dates = sorted(segments['date'].dropna().unique())
    clear_saved_charts()

    def filter_by_area(source_df, area_key):
        if area_key is None: return source_df.copy()
        return source_df[source_df['area_key'] == area_key].copy()

    combined_sections = []; overall_by_date = {}
    for the_date in all_dates:
        df_day = segments[segments['date'] == the_date].copy()
        if df_day.empty: continue
        file_date_mode = df_day['file_date'].dropna().astype(str)
        display_date = pd.to_datetime(file_date_mode.mode().iloc[0]).date() if not file_date_mode.empty else pd.to_datetime(the_date).date()
        date_str = format_file_date_for_title(display_date)
        print(f"\n🎨 Дата (из файла): {date_str}")
        ws = datetime.combine(display_date, dtime(6,0))
        we = datetime.combine(display_date, dtime(23,0))
        df_day_8_22 = df_day[(pd.to_datetime(df_day['start']) >= ws) & (pd.to_datetime(df_day['start']) < we)].copy()

        areas_today = (df_day['area_key'].fillna('').astype(str).str.strip().unique().tolist())
        for ak in areas_today:
            df_area = filter_by_area(df_day, ak if ak else '')
            area_label = df_area['area'].dropna().astype(str).str.strip().unique()
            area_label = area_label[0] if len(area_label)>0 else "Без участка"
            if df_area.empty: continue
            for emp, g in df_area.groupby('employee'):
                fig, panel_html = create_user_timeline_chart_window(g, tag_desc_map=tag_desc_map, row_height=row_height)
                if not fig: continue
                fig.update_layout(title_text=f"{emp} — {date_str}")
                fig_div = fig_to_div(fig, fig.layout.title.text)
                combined_sections.append({'heading': f"{date_str}", 'subheading': f"{emp}",
                                          'panel_html': panel_html, 'fig_div': fig_div})
                LAST_EXPORTS.append({'fig': fig, 'date_str': date_str, 'area_suffix': area_label, 'employee': emp})
        overall_by_date[the_date] = df_day_8_22.copy()

    if all_dates:
        if len(all_dates)==1:
            page_title = f"AA_BLE Таймлайны (6–23) — {date_str}"
            filename = f"AA_BLE Таймлайны 6-23 {date_str}.html"
        else:
            min_d = format_file_date_for_title(pd.to_datetime(min(all_dates)).date())
            max_d = format_file_date_for_title(pd.to_datetime(max(all_dates)).date())
            page_title = f"AA_BLE Таймлайны (6–23) — период {min_d}–{max_d}"
            filename = f"AA_BLE Таймлайны 6-23 {min_d}—{max_d}.html"
        export_combined_html(combined_sections, filename, page_title)

    export_overall_excel_period(overall_by_date)
    print("\n✅ ГОТОВО! Окно 6:00–23:00, поминутный учёт; справа — отдельное окно со временем по ЗОНАМ; на диаграмме метки, окрашенные в оттенки своей зоны.")
    print("Повторная выгрузка отдельных графиков: redownload_last_charts().")

print("=" * 100)
print("🏢 AA_BLE ТАЙМЛАЙНЫ (6:00–23:00, поминутно; справа — окно «Время по зонам»)")
print("=" * 100)
print("🧭 Зоны: 1 зелёные (проведения работ), 4 красные (курилки), 10 синие (выдача WW), 13 оранжевые (КПП), 0 серые, прочее — серые.")
print("📅 Дата в отчёте берётся из имени файла (YYYY-MM-DD и др. разделители → DD-MM-YYYY).")
print("🚀 Запуск: create_aable_presence_analysis()")
print("=" * 100)
print()
print("✅ ГОТОВО! Запустите create_aable_presence_analysis() для начала.")
